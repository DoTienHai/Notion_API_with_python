import json
import os
import pandas as pd
from openpyxl import Workbook
from Config import *
import numpy as np

def convert_data_in_cell_to_json(string_data):
    if isinstance(string_data, str):
        string_data = string_data.replace("'",'"')
        string_data = string_data.replace('None', 'null')
        string_data = string_data.replace('False', 'false')
        string_data = string_data.replace('True', 'true')
        json_data = json.loads(string_data)
        return json_data
    else:
        return ""

def extract_id(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        return item[0]["id"]
    else:
        return ""

def extract_number(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        return item[0]["number"]
    else:
        return ""

def extract_plain_text(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        return item[0]["plain_text"]
    else:
        return ""

def extract_text_content(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        return item[0]["text"]["content"]
    else:
        return ""

def extract_select_name(item):
    item = convert_data_in_cell_to_json(item)
    if (len(item) > 0):
        if (item[0]["select"]):
            return item[0]["select"]["name"]
        else:
            return ""
    else:
        return ""

def extract_relation_id(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        if len(item[0]["relation"]):
            return item[0]["relation"][0]["id"]
        else:
                return ""
    else:
        return 0

def extract_multi_select(item):
    items = convert_data_in_cell_to_json(item)
    ret = []
    if len(items) > 0:
        for i in items:
            ret.append(i["name"])
        return ",".join(ret)
    else:
        return ""

def convert_json_to_excel():
    json_file_paths = []
    # Duyệt từng tệp và thư mục trong thư mục hiện tại
    for root, dirs, files in os.walk(os.path.join(notion_data_folder)):
        for file in files:
            # Kiểm tra nếu tệp có đuôi là suffix
            if file.endswith('.json'):
                json_file_paths.append(os.path.join(root, file))
    if json_file_paths:
        for json_file_path in json_file_paths:
            with open(json_file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            # Chuyển đổi thành DataFrame
            df = pd.json_normalize(data)
            fileName = os.path.basename(json_file_path)
            part_split = fileName.split(".")
            file_path = os.path.join(notion_data_folder, f"{part_split[0]}.xlsx")
            df.to_excel(file_path,sheet_name=f"{part_split[0]}", index=False)
            print(f"Convert {fileName} to excel!")
    else:
        print("Json file not found!")



def collect_ho_so_nhan_su():
    excel_file_path = os.path.join(notion_data_folder, "Hồ sơ nhân sự.xlsx")
    data_ho_so_nhan_su = pd.read_excel(excel_file_path)
    data_ho_so_nhan_su = data_ho_so_nhan_su[["id", "properties.Mã nhân viên.unique_id.prefix", "properties.Mã nhân viên.unique_id.number",
                                             "properties.Họ và tên.title", "properties.Quê quán.rich_text",                                          
                                             "properties.SĐT.phone_number", "properties.Email.email", 
                                             "properties.Ngày sinh.date.start", "properties.Chức vụ.select.name", 
                                             "properties.Ngày bắt đầu làm việc.date.start", "properties.Cơ sở.select.name", "properties.Hình thức làm việc.select.name",
                                             "properties.Phân cấp.select.name", "properties.Doanh số.number",
                                             "properties.KPI.number","properties.Tỉ lệ đạt KPI.formula.number"]]
    data_ho_so_nhan_su = data_ho_so_nhan_su.rename(columns={"id":"notion id", "properties.Mã nhân viên.unique_id.prefix":"Tiền tố", "properties.Mã nhân viên.unique_id.number":"Mã nhân viên",
                                             "properties.Họ và tên.title":"Họ và tên", "properties.Quê quán.rich_text":"Quê quán",                                          
                                             "properties.SĐT.phone_number":"SĐT", "properties.Email.email":"Email", 
                                             "properties.Ngày sinh.date.start":"Ngày sinh", "properties.Chức vụ.select.name":"Chức vụ", 
                                             "properties.Ngày bắt đầu làm việc.date.start":"Ngày bắt đầu làm việc", "properties.Cơ sở.select.name":"Cơ sở", 
                                             "properties.Phân cấp.select.name":"Phân cấp", "properties.Hình thức làm việc.select.name" : "Hình thức làm việc", 
                                             "properties.KPI.number":"KPI", "properties.Doanh số.number" : "Doanh số",
                                             "properties.Tỉ lệ đạt KPI.formula.number" : "Tỉ lệ đạt KPI"})

    data_ho_so_nhan_su['Họ và tên'] = data_ho_so_nhan_su['Họ và tên'].apply(extract_text_content)
    data_ho_so_nhan_su["Quê quán"] = data_ho_so_nhan_su["Quê quán"].apply(extract_text_content)
    # Xử lý cột SĐT
    def extract_SDT(item):
        if(item is None or np.isnan(item)):
            return ""
        else:
            return f"0{round(item)}"
    data_ho_so_nhan_su["SĐT"] = data_ho_so_nhan_su["SĐT"].apply(extract_SDT).astype(str)
    return data_ho_so_nhan_su.sort_values("Mã nhân viên")

def collect_thong_tin_khach_hang():
    excel_file_path = os.path.join(notion_data_folder, "Thông tin khách hàng.xlsx")
    data_thong_tin_khach_hang = pd.read_excel(excel_file_path)
    data_thong_tin_khach_hang = data_thong_tin_khach_hang[["id", "properties.Mã khách hàng.unique_id.prefix", 
                                           "properties.Mã khách hàng.unique_id.number", "properties.Họ và tên.title", 
                                           "properties.Cơ sở.rollup.array", "properties.CCCD.rich_text",  "properties.SĐT.phone_number", 
                                           "properties.Link Facebook.url", "properties.Địa chỉ.rich_text", "properties.Tích lũy.rollup.number",
                                           "properties.Dư nợ.rollup.number"]]     
    data_thong_tin_khach_hang = data_thong_tin_khach_hang.rename(columns={"id":"notion id", "properties.Mã khách hàng.unique_id.prefix":"Tiền tố", 
                                           "properties.Mã khách hàng.unique_id.number":"Mã khách hàng", "properties.Họ và tên.title":"Họ và tên", 
                                           "properties.CCCD.rich_text":"CCCD",  "properties.SĐT.phone_number":"SĐT", 
                                           "properties.Link Facebook.url":"Facebook", "properties.Địa chỉ.rich_text":"Địa chỉ",
                                            "properties.Cơ sở.rollup.array" : "Cơ sở", "properties.Tích lũy.rollup.number" : "Tích lũy",
                                            "properties.Dư nợ.rollup.number" : "Dư nợ"})

    data_thong_tin_khach_hang["Họ và tên"] = data_thong_tin_khach_hang["Họ và tên"].apply(extract_plain_text)
    data_thong_tin_khach_hang["CCCD"] = data_thong_tin_khach_hang["CCCD"].apply(extract_text_content)
    data_thong_tin_khach_hang["Địa chỉ"] = data_thong_tin_khach_hang["Địa chỉ"].apply(extract_text_content)
    data_thong_tin_khach_hang["Cơ sở"] = data_thong_tin_khach_hang["Cơ sở"].apply(extract_select_name)
    return data_thong_tin_khach_hang

def collect_danh_muc_dich_vu():
    excel_file_path = os.path.join(notion_data_folder, "Danh mục dịch vụ.xlsx")
    data_danh_muc_dich_vu = pd.read_excel(excel_file_path)
    data_danh_muc_dich_vu = data_danh_muc_dich_vu[["id", "properties.Tên dịch vụ.title", "properties.Nhóm dịch vụ.select.name",
                                           "properties.Số ca.rollup.number", "properties.Công phụ phẫu 1.number",
                                           "properties.Công phụ phẫu 2.number"]]
                                           
    data_danh_muc_dich_vu = data_danh_muc_dich_vu.rename(columns={"id":"notion id", "properties.Tên dịch vụ.title": "Tên dịch vụ", 
                                                          "properties.Nhóm dịch vụ.select.name":"Nhóm dịch vụ",
                                                            "properties.Số ca.rollup.number":"Số ca", 
                                                            "properties.Công phụ phẫu 1.number":"Công phụ phẫu 1",
                                                            "properties.Công phụ phẫu 2.number":"Công phụ phẫu 2"})
    data_danh_muc_dich_vu['Tên dịch vụ'] = data_danh_muc_dich_vu['Tên dịch vụ'].apply(extract_text_content)

    return data_danh_muc_dich_vu.sort_values("Tên dịch vụ")

def collect_chi_tieu():
    excel_file_path = os.path.join(notion_data_folder, "Chi tiêu.xlsx")
    data_chi_tieu = pd.read_excel(excel_file_path)
    data_chi_tieu = data_chi_tieu[["id", "properties.Auto mã chi tiêu.unique_id.prefix", "properties.Auto mã chi tiêu.unique_id.number",
                               "properties.Ngày chi.date.start", "properties.Cơ sở.select.name","properties.Phân loại.select.name", "properties.Nhân viên xác nhận.relation",
                               "properties.Lượng chi.number", "properties.Người Nhận/Ứng.relation", "properties.Ghi chú.rich_text"]]
    data_chi_tieu = data_chi_tieu.rename(columns={"id":"notion id", "properties.Ngày chi.date.start":"Ngày chi", "properties.Nhân viên xác nhận.relation":"id nhân viên xác nhận",
                               "properties.Cơ sở.select.name":"Cơ sở", "properties.Người Nhận/Ứng.relation":"id người nhận/ứng", 
                               "properties.Phân loại.select.name":"Phân loại", "properties.Auto mã chi tiêu.unique_id.prefix":"Tiền tố", 
                               "properties.Auto mã chi tiêu.unique_id.number":"Mã chi tiêu", "properties.Lượng chi.number":"Lượng chi", "properties.Ghi chú.rich_text" : "Ghi chú"})

    data_chi_tieu["id nhân viên xác nhận"] = data_chi_tieu["id nhân viên xác nhận"].apply(extract_id)
    data_chi_tieu["id người nhận/ứng"] = data_chi_tieu["id người nhận/ứng"].apply(extract_id)
    data_chi_tieu["Ghi chú"] = data_chi_tieu["Ghi chú"].apply(extract_text_content)
    data_chi_tieu = pd.merge(data_chi_tieu, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id nhân viên xác nhận", right_on="notion id", how="left")
    data_chi_tieu = data_chi_tieu.rename(columns={"Họ và tên":"Nhân viên xác nhận"})
    data_chi_tieu = pd.merge(data_chi_tieu, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id người nhận/ứng", right_on="notion id", how="left")
    data_chi_tieu = data_chi_tieu.rename(columns={"Họ và tên":"Người nhận/ứng"})
    data_chi_tieu = data_chi_tieu.drop(columns=["notion id"])
    data_chi_tieu = data_chi_tieu.drop(columns=["notion id_y"])
    data_chi_tieu = data_chi_tieu.rename(columns={"notion id_x":"notion id"})
    return data_chi_tieu.sort_values("Mã chi tiêu")

def calculate_ti_le_chiet_khau(nhom_dich_vu, don_gia):
    if nhom_dich_vu:
        ref_chiet_khau = pd.read_excel("Ref tính lương.xlsx", sheet_name="Chiết khấu")
        cols = ref_chiet_khau.columns.to_list()
        cols.remove("Nhóm dịch vụ")

        temp = "0"
        for col in cols:
            if int(col) <= don_gia and int(col) > int(temp):
                temp = col
        
        ti_le_chiet_khau = ref_chiet_khau[ref_chiet_khau["Nhóm dịch vụ"] == nhom_dich_vu][temp]
        return ti_le_chiet_khau.values[0]
    else:
        return 0

def collect_doanh_thu_he_thong():
    excel_file_path = os.path.join(notion_data_folder, "Doanh thu HỆ THỐNG.xlsx")
    data_doanh_thu_he_thong = pd.read_excel(excel_file_path)
    data_doanh_thu_he_thong = data_doanh_thu_he_thong[["id", "properties.Auto mã dịch vụ.unique_id.prefix", 
                                               "properties.Auto mã dịch vụ.unique_id.number", "properties.Ngày thực hiện.date.start", 
                                               "properties.Cơ sở.select.name", "properties.Loại hình dịch vụ.relation", "properties.Nhóm dịch vụ.rollup.array",
                                               "properties.Khách hàng.relation", "properties.Nguồn khách.select.name",
                                               "properties.Sale chính.relation", "properties.Đơn giá gốc.number",
                                               "properties.Sale phụ.relation","properties.Upsale.number","properties.Đơn giá.formula.number",
                                                "properties.Thanh toán lần đầu.number", "properties.Trả sau.rollup.number",
                                                "properties.Đã thanh toán.formula.number", "properties.Dư nợ.formula.number",
                                               "properties.Bác sĩ 1.relation",  "properties.Bác sĩ 2.relation",
                                                "properties.Phụ phẫu 1.relation", "properties.Phụ phẫu 2.relation", "properties.Công phụ phẫu 1.rollup.array",
                                                "properties.Công phụ phẫu 2.rollup.array"
                                               ]]
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"id":"notion id", "properties.Phụ phẫu 1.relation":"id phụ phẫu 1", 
                                               "properties.Khách hàng.relation":"id khách hàng", "properties.Auto mã dịch vụ.unique_id.prefix":"Tiền tố", 
                                               "properties.Auto mã dịch vụ.unique_id.number":"Mã dịch vụ", "properties.Đã thanh toán.formula.number":"Đã thanh toán", 
                                               "properties.Cơ sở.select.name":"Cơ sở", "properties.Bác sĩ 1.relation":"id bác sĩ 1", 
                                               "properties.Thanh toán lần đầu.number":"Thanh toán lần đầu", "properties.Đơn giá.formula.number":"Đơn giá", 
                                               "properties.Dư nợ.formula.number":"Dư nợ", "properties.Phụ phẫu 2.relation":"id phụ phẫu 2", 
                                               "properties.Bác sĩ 2.relation":"id bác sĩ 2", "properties.Sale chính.relation":"id sale chính", "properties.Đơn giá gốc.number":"Đơn giá gốc",
                                               "properties.Sale phụ.relation":"id sale phụ","properties.Upsale.number":"Upsale",
                                               "properties.Nguồn khách.select.name":"Nguồn khách", "properties.Trả sau.rollup.number":"Trả sau", 
                                               "properties.Ngày thực hiện.date.start":"Ngày thực hiện", "properties.Loại hình dịch vụ.relation":"id loại hình dịch vụ", 
                                               "properties.Công phụ phẫu 1.rollup.array":"Công phụ phẫu 1", "properties.Công phụ phẫu 2.rollup.array":"Công phụ phẫu 2", 
                                               "properties.Nhóm dịch vụ.rollup.array" : "Nhóm dịch vụ"})
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.fillna("")


    data_doanh_thu_he_thong["Nhóm dịch vụ"] = data_doanh_thu_he_thong["Nhóm dịch vụ"].apply(extract_select_name)
    data_doanh_thu_he_thong["id khách hàng"] = data_doanh_thu_he_thong["id khách hàng"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_thong_tin_khach_hang()[["notion id", "Họ và tên"]], left_on="id khách hàng", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"Họ và tên":"Khách hàng"})
    data_doanh_thu_he_thong["id sale chính"] = data_doanh_thu_he_thong["id sale chính"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id sale chính", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"Họ và tên":"Sale chính"})
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id"])
    data_doanh_thu_he_thong["id sale phụ"] = data_doanh_thu_he_thong["id sale phụ"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id sale phụ", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"Họ và tên":"Sale phụ"})
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id"])
    data_doanh_thu_he_thong["id bác sĩ 1"] = data_doanh_thu_he_thong["id bác sĩ 1"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id bác sĩ 1", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"Họ và tên":"Bác sĩ 1"})
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id"])
    data_doanh_thu_he_thong["id bác sĩ 2"] = data_doanh_thu_he_thong["id bác sĩ 2"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id bác sĩ 2", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"Họ và tên":"Bác sĩ 2"})
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id"])
    data_doanh_thu_he_thong["id phụ phẫu 1"] = data_doanh_thu_he_thong["id phụ phẫu 1"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id phụ phẫu 1", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"Họ và tên":"Phụ phẫu 1"})
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id"])
    data_doanh_thu_he_thong["id phụ phẫu 2"] = data_doanh_thu_he_thong["id phụ phẫu 2"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id phụ phẫu 2", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"Họ và tên":"Phụ phẫu 2"})
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id"])
    data_doanh_thu_he_thong["id loại hình dịch vụ"] = data_doanh_thu_he_thong["id loại hình dịch vụ"].apply(extract_id)
    data_doanh_thu_he_thong = pd.merge(data_doanh_thu_he_thong, collect_danh_muc_dich_vu()[["notion id", "Tên dịch vụ"]], left_on="id loại hình dịch vụ", right_on="notion id", how="left")
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id"])
    
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.drop(columns=["notion id_y"])
    data_doanh_thu_he_thong = data_doanh_thu_he_thong.rename(columns={"notion id_x":"notion id"})
    data_doanh_thu_he_thong["Công phụ phẫu 1"] =  data_doanh_thu_he_thong["Công phụ phẫu 1"].apply(extract_number)
    data_doanh_thu_he_thong["Công phụ phẫu 2"] =  data_doanh_thu_he_thong["Công phụ phẫu 2"].apply(extract_number)
    # Tính tỉ lẹ chiết khấu bác sĩ 
    for index, row in data_doanh_thu_he_thong.iterrows():
        if pd.notna(row["Bác sĩ 2"]):
            data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu bác sĩ 1"] = 0.06
            data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu bác sĩ 2"] = 0.06
        else:
            data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu bác sĩ 1"] = 0.1
            data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu bác sĩ 2"] = 0
            # Tỉ lệ chiết khấu đơn 1 bác sĩ của NV-5 Nguyễn Hoàng Yến Quyên 75046948-a198-4627-89b3-3bbf5967526b
            if (row["id bác sĩ 1"] == "75046948-a198-4627-89b3-3bbf5967526b"):
                if ((row["Nhóm dịch vụ"] == "Tiêm") or (row["Nhóm dịch vụ"] == "Tiểu phẫu")) and (row["Cơ sở"] == "CẦN THƠ"):
                    data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu bác sĩ 1"] = 0.08
    data_doanh_thu_he_thong["Chiết khấu bác sĩ 1"] = data_doanh_thu_he_thong["Tỉ lệ chiết khấu bác sĩ 1"]*data_doanh_thu_he_thong["Đã thanh toán"]
    data_doanh_thu_he_thong["Chiết khấu bác sĩ 2"] = data_doanh_thu_he_thong["Tỉ lệ chiết khấu bác sĩ 2"]*data_doanh_thu_he_thong["Đã thanh toán"]
    # Tính tỉ lệ chiết khấu sale
    data_doanh_thu_he_thong["Tỉ lệ chiết khấu sale chính"] = 0 
    data_doanh_thu_he_thong["Tỉ lệ chiết khấu sale phụ"] = 0 
    data_doanh_thu_he_thong["Chiết khấu sale chính"] = 0 
    data_doanh_thu_he_thong["Chiết khấu sale phụ"] = 0 
    for index, row in data_doanh_thu_he_thong.iterrows():
        if row["Nguồn khách"] == "Cá nhân":
            ti_le_chieu_khau = calculate_ti_le_chiet_khau(row["Nhóm dịch vụ"], row["Đơn giá"])
            # Tỉ lệ chiết khấu trường hợp Lê Đình Hâu NV-10 c463b1a9-4fb2-4258-87a7-44193ba02405
            if (row["id sale chính"] == "c463b1a9-4fb2-4258-87a7-44193ba02405") and ((row["Cơ sở"] == "CẦN THƠ") or (row["Cơ sở"] == "LONG XUYÊN")):
                ti_le_chieu_khau = 0.15      
            data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu sale chính"] = ti_le_chieu_khau
            
            if pd.notna(row["Sale phụ"]):
                data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu sale phụ"] = 0.04
        elif row["Nguồn khách"] == "CTV":
            if pd.notna(row["Sale phụ"]):
                data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu sale phụ"] = 0.02
        elif row["Nguồn khách"] == "Khách cũ" or row["Nguồn khách"] == "Khách cũ giới thiệu":
            data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu sale chính"] = 0.1
            if pd.notna(row["Sale phụ"]):
                data_doanh_thu_he_thong.at[index , "Tỉ lệ chiết khấu sale phụ"] = 0.04        
        elif row["Nguồn khách"] == "Page":
            pass
        elif row["Nguồn khách"] == "Khách cửa hàng":
            data_doanh_thu_he_thong.at[index, "Tỉ lệ chiết khấu sale chính"] = 0.04
            if pd.notna(row["Sale phụ"]):
                data_doanh_thu_he_thong.at[index, "Tỉ lệ chiết khấu sale phụ"] = 0.02

    # Tính chiết khấu
    for index, row in data_doanh_thu_he_thong.iterrows():
        don_gia_goc = row["Đơn giá gốc"]
        da_thanh_toan = row["Đã thanh toán"]
        if (isinstance(don_gia_goc, float) or isinstance(don_gia_goc, int)) and (isinstance(da_thanh_toan, float) or isinstance(da_thanh_toan, int)):
            if (da_thanh_toan >= don_gia_goc):
                data_doanh_thu_he_thong.at[index, "Chiết khấu sale chính"] = don_gia_goc*row["Tỉ lệ chiết khấu sale chính"] + (da_thanh_toan-don_gia_goc)*(row["Tỉ lệ chiết khấu sale chính"]-row["Tỉ lệ chiết khấu sale phụ"])
                data_doanh_thu_he_thong.at[index, "Chiết khấu sale phụ"] = (da_thanh_toan-don_gia_goc)*row["Tỉ lệ chiết khấu sale phụ"]
            else:
                data_doanh_thu_he_thong.at[index, "Chiết khấu sale chính"] = da_thanh_toan*row["Tỉ lệ chiết khấu sale chính"] 
    
    return data_doanh_thu_he_thong.sort_values("Mã dịch vụ")


def collect_danh_sach_thu_no(data_doanh_thu):
    excel_file_path = os.path.join(notion_data_folder, "Danh sách thu nợ.xlsx")
    data_thu_no = pd.read_excel(excel_file_path)
    data_thu_no = data_thu_no[["id", "properties.Gen mã đơn.unique_id.prefix", "properties.Gen mã đơn.unique_id.number",
                           "properties.Ngày thu.date.start", "properties.Cơ sở.rollup.array",
                           "properties.Đơn nợ.relation", "properties.Lượng thu.number"]]
    data_thu_no = data_thu_no.rename(columns={"id":"notion id", "properties.Gen mã đơn.unique_id.prefix":"Tiền tố", "properties.Gen mã đơn.unique_id.number": "Mã đơn thu nợ",
                           "properties.Ngày thu.date.start":"Ngày thu", "properties.Cơ sở.rollup.array":"Cơ sở",
                           "properties.Đơn nợ.relation":"id đơn nợ", "properties.Lượng thu.number":"Lượng thu"})

    data_thu_no["id đơn nợ"] = data_thu_no["id đơn nợ"].apply(extract_id)
    data_thu_no["Cơ sở"] = data_thu_no["Cơ sở"].apply(extract_select_name)

    data_doanh_thu = data_doanh_thu[["notion id", "Mã dịch vụ", "Khách hàng", "Nguồn khách", "Sale chính","Đơn giá gốc", "Sale phụ", "Upsale", 
                                    "Tên dịch vụ", "Bác sĩ 1", "Bác sĩ 2", "Thanh toán lần đầu", "Đã thanh toán", "Đơn giá", 
                                    "Tỉ lệ chiết khấu sale chính", "Tỉ lệ chiết khấu sale phụ", "Ngày thực hiện", 
                                    "id sale chính", "id sale phụ", "id bác sĩ 1", "id bác sĩ 2", 
                                    "Tỉ lệ chiết khấu bác sĩ 1", "Tỉ lệ chiết khấu bác sĩ 2"]]
    data_thu_no = pd.merge(data_thu_no, data_doanh_thu, left_on="id đơn nợ", right_on="notion id", how="left")
    data_thu_no = data_thu_no.rename(columns={"Mã dịch vụ":"Đơn nợ"})
    def format_don_no(item):
        item = '{:.0f}'.format(item)
        return f"HD-LUXURY-{item}"
    data_thu_no["Đơn nợ"] = data_thu_no["Đơn nợ"].apply(format_don_no)

    data_thu_no = data_thu_no.drop(columns=["notion id_y"])
    data_thu_no = data_thu_no.rename(columns={"notion id_x":"notion id"})
    
    data_thu_no["Chiết khấu bác sĩ 1"] = 0
    data_thu_no["Chiết khấu bác sĩ 2"] = 0
    data_thu_no["Chiết khấu sale chính"] = 0
    data_thu_no["Chiết khấu sale phụ"] = 0
    for index, row in data_thu_no.iterrows():
        luong_thu = row["Lượng thu"]
        don_gia_goc = row["Đơn giá gốc"]
        da_thanh_toan = row["Đã thanh toán"]

        data_thu_no.at[index, "Chiết khấu bác sĩ 1"] = row["Tỉ lệ chiết khấu bác sĩ 1"]*luong_thu
        data_thu_no.at[index, "Chiết khấu bác sĩ 2"] = row["Tỉ lệ chiết khấu bác sĩ 2"]*luong_thu

        if row["Nguồn khách"] == "Cá nhân" or row["Nguồn khách"] == "Khách cửa hàng" or row["Nguồn khách"] == "Khách cũ" or row["Nguồn khách"] == "Khách cũ giới thiệu":
            if (da_thanh_toan >= don_gia_goc):
                da_thanh_toan_truoc_thu_no = da_thanh_toan - luong_thu
                if da_thanh_toan_truoc_thu_no <= don_gia_goc:
                    phan_sale_phu = da_thanh_toan - don_gia_goc
                    phan_sale_chinh = luong_thu - phan_sale_phu
                    data_thu_no.at[index, "Chiết khấu sale chính"] = phan_sale_chinh*row["Tỉ lệ chiết khấu sale chính"] + phan_sale_phu*(row["Tỉ lệ chiết khấu sale chính"]-row["Tỉ lệ chiết khấu sale phụ"])
                    data_thu_no.at[index, "Chiết khấu sale phụ"] = phan_sale_phu*row["Tỉ lệ chiết khấu sale phụ"]
                else:
                    data_thu_no.at[index, "Chiết khấu sale chính"] = luong_thu*row["Tỉ lệ chiết khấu sale chính"] + luong_thu*(row["Tỉ lệ chiết khấu sale chính"]-row["Tỉ lệ chiết khấu sale phụ"])

            else:
                data_thu_no.at[index, "Chiết khấu sale chính"] = luong_thu*row["Tỉ lệ chiết khấu sale chính"]
        elif row["Nguồn khách"] == "CTV":
                data_thu_no.at[index, "Chiết khấu sale phụ"] = luong_thu*row["Tỉ lệ chiết khấu sale phụ"]
        elif row["Nguồn khách"] == "Page":
            pass
    
    return data_thu_no.sort_values("Mã đơn thu nợ")

def collect_data_cham_cong_he_thong():
    excel_file_path = os.path.join(notion_data_folder, "Chấm công HỆ THỐNG.xlsx")
    data_cham_cong = pd.read_excel(excel_file_path)
    raw_columns = ["id","properties.Nhân sự.relation", "properties.Cơ sở.rollup.array", "properties.Tổng công.formula.number"]
    new_columns = ["notion id", "id nhân sự", "Cơ sở", "Tổng công"]
    for key,value in cham_cong_ref.items():
        raw_columns.append(f"properties.{key}.number")
        new_columns.append(f"{key}")
    for location in location_list:
        if location != "HỆ THỐNG":
            raw_columns.append(f"properties.Tổng công tại {location}.number")
            new_columns.append(f"Tổng công tại {location}")
    columns_dict = dict(zip(raw_columns, new_columns))
    data_cham_cong = data_cham_cong[raw_columns]
    data_cham_cong = data_cham_cong.rename(columns=columns_dict)

    data_cham_cong["id nhân sự"] = data_cham_cong["id nhân sự"].apply(extract_id)
    data_cham_cong["Cơ sở"] = data_cham_cong["Cơ sở"].apply(extract_select_name)
    data_cham_cong = pd.merge(data_cham_cong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id nhân sự", right_on="notion id", how="left")
    
    data_cham_cong = data_cham_cong.drop(columns=["notion id_y"])
    data_cham_cong = data_cham_cong.rename(columns={"notion id_x":"notion id"})
    return data_cham_cong

def collect_data_cham_cong_co_so(location):
    excel_file_path = os.path.join(notion_data_folder, f"Chấm công {location}.xlsx")
    data_cham_cong = pd.read_excel(excel_file_path)
    select_name_columns = ["id","properties.Nhân sự.relation", "properties.Cơ sở.rollup.array"]
    select_columns = ["id","properties.Nhân sự.relation", "properties.Cơ sở.rollup.array"]
    raw_columns = []
    new_columns = ["notion id", "id nhân sự", "Cơ sở"]
    for i in range(31):
        select_name_columns.append(f"properties.Ngày {i+1}.select.name")
        select_columns.append(f"properties.Ngày {i+1}.select")
        new_columns.append(f"Ngày {i+1}")
    col_data = data_cham_cong.columns.to_list()
    # nếu cột select chưa được chọn sẽ không có cột ...select.name sẽ chỉ chọn cột ...select
    for item in select_name_columns:
        if item in col_data:
            raw_columns.append(item)
        else:
            raw_columns.append(select_columns[select_name_columns.index(item)])
    columns_dict = dict(zip(raw_columns, new_columns))
    data_cham_cong = data_cham_cong[raw_columns]
    data_cham_cong = data_cham_cong.rename(columns=columns_dict)


    data_cham_cong["id nhân sự"] = data_cham_cong["id nhân sự"].apply(extract_id)
    data_cham_cong["Cơ sở"] = data_cham_cong["Cơ sở"].apply(extract_select_name)
    data_cham_cong = pd.merge(data_cham_cong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id nhân sự", right_on="notion id", how="left")
    
    data_cham_cong = data_cham_cong.drop(columns=["notion id_y"])
    data_cham_cong = data_cham_cong.rename(columns={"notion id_x":"notion id"})
    return data_cham_cong

def collect_data_thuong_phat():
    excel_file_path = os.path.join(notion_data_folder, "Thưởng phạt.xlsx")
    data_thuong_phat = pd.read_excel(excel_file_path)
    data_thuong_phat = data_thuong_phat[["id", "properties.Auto mã.unique_id.prefix", "properties.Auto mã.unique_id.number",
                                         "properties.Nhân sự.relation", "properties.Cơ sở.rollup.array",
                                         "properties.Loại.select.name", "properties.Lượng thưởng phạt.number", 
                                          "properties.Ngày phát sinh.date.start", "properties.Lí do.rich_text" ]]
    data_thuong_phat = data_thuong_phat.rename(columns={"id":"notion id", "properties.Auto mã.unique_id.prefix" : "Tiền tố", 
                                                        "properties.Auto mã.unique_id.number" : "Mã thưởng phạt",
                                                        "properties.Nhân sự.relation" : "id nhân sự", "properties.Cơ sở.rollup.array" : "Cơ sở",
                                                        "properties.Loại.select.name" : "Loại", "properties.Lượng thưởng phạt.number" : "Lượng thưởng phạt", 
                                                        "properties.Ngày phát sinh.date.start" : "Ngày phát sinh", "properties.Lí do.rich_text" : "Lí do" })
    data_thuong_phat["id nhân sự"] = data_thuong_phat["id nhân sự"].apply(extract_id)
    data_thuong_phat = pd.merge(data_thuong_phat, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id nhân sự", right_on="notion id", how="left")
    data_thuong_phat = data_thuong_phat.drop(columns=["notion id_y"])
    data_thuong_phat = data_thuong_phat.rename(columns={"notion id_x":"notion id"})
    data_thuong_phat["Cơ sở"] = data_thuong_phat["Cơ sở"].apply(extract_select_name)
    data_thuong_phat["Lí do"] = data_thuong_phat["Lí do"].apply(extract_text_content)
    return data_thuong_phat


def collect_data():
    convert_json_to_excel()
    # Kiểm tra xem file Excel đã tồn tại hay chưa
    excel_file_path = file_all_notion_data
    if os.path.exists(excel_file_path):
        # Nếu đã tồn tại, xóa file cũ đi
        try:
            os.remove(excel_file_path)
            print(f"Đã xóa file Excel cũ '{excel_file_path}'")
        except Exception as e:
            print(f"Lỗi khi xóa file Excel cũ: {e}")

    # Tạo workbook mới
    wb = Workbook()
    # Tạo sheet hồ sơ nhân sự
    ws1 = wb.active
    ws1.title = 'Hồ sơ nhân sự'
    writeDataframeToSheet(ws1, collect_ho_so_nhan_su())
    # Tạo sheet danh sách khách hàng
    ws2 = wb.create_sheet("Danh sách khách hàng")
    writeDataframeToSheet(ws2, collect_thong_tin_khach_hang())
    # Tạo sheet danh mục dịch vụ
    ws3 = wb.create_sheet("Danh mục dịch vụ")
    writeDataframeToSheet(ws3, collect_danh_muc_dich_vu())
    # Tạo sheet chi tiêu
    ws4 = wb.create_sheet(title='Chi tiêu')
    writeDataframeToSheet(ws4, collect_chi_tieu())
    # Tạo sheet doanh thu HỆ THỐNG
    ws5 = wb.create_sheet(title="Doanh thu HỆ THỐNG")
    data_doanh_thu = collect_doanh_thu_he_thong()
    writeDataframeToSheet(ws5, data_doanh_thu)
    # Tạo sheet Thu nợ
    ws6 = wb.create_sheet(title="Thu nợ")
    writeDataframeToSheet(ws6, collect_danh_sach_thu_no(data_doanh_thu))
    # Tạo sheet chấm công HỆ THỐNG
    ws7 = wb.create_sheet(title=f"Chấm công HỆ THỐNG")
    writeDataframeToSheet(ws7, collect_data_cham_cong_he_thong())
    # Tạo sheet chấm công các cơ sở
    for location in location_list:
        if location != "HỆ THỐNG": 
            ws_cham_cong = wb.create_sheet(title=f"Chấm công {location}")
            writeDataframeToSheet(ws_cham_cong, collect_data_cham_cong_co_so(location))
    # Tạo sheet thưởng phạt
    ws8 = wb.create_sheet("Thưởng phạt")
    writeDataframeToSheet(ws8, collect_data_thuong_phat())

    # Lưu workbook vào file Excel
    try:
        wb.save(excel_file_path)
        print(f"Đã tạo file Excel mới '{excel_file_path}' thành công")
    except Exception as e:
        print(f"Lỗi khi tạo file Excel mới: {e}")


    