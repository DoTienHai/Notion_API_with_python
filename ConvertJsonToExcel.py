import json
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from Utils import *
import numpy as np

def convert_data_in_cell_to_json(string_data):
    if isinstance(string_data, str):
        string_data = string_data.replace("'",'"')
        string_data = string_data.replace('None', 'null')
        string_data = string_data.replace('False', 'false')
        string_data = string_data.replace('True', 'true')
        json_data = json.loads(string_data)
        return json_data
    else:
        return ""

def extract_id(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        return item[0]["id"]
    else:
        return ""

def extract_plain_text(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        return item[0]["plain_text"]
    else:
        return ""

def extract_text_content(item):
    item = convert_data_in_cell_to_json(item)
    if len(item) > 0:
        return item[0]["text"]["content"]
    else:
        return ""

def extract_select_name(item):
    item = convert_data_in_cell_to_json(item)
    if (len(item) > 0):
        if (item[0]["select"]):
            return item[0]["select"]["name"]
        else:
            return ""
    else:
        return ""

def convert_json_to_excel(directory = "./", suffix='.json'):
    jsonFilePath = []
    # Duyệt từng tệp và thư mục trong thư mục hiện tại
    for root, dirs, files in os.walk(directory):
        for file in files:
            # Kiểm tra nếu tệp có đuôi là suffix
            if file.endswith(suffix):
                jsonFilePath.append(os.path.join(root, file))
    if jsonFilePath:
        for jsonFilePath in jsonFilePath:
            with open(jsonFilePath, 'r', encoding='utf-8') as file:
                data = json.load(file)
            # Chuyển đổi thành DataFrame
            df = pd.json_normalize(data)
            fileName = os.path.basename(jsonFilePath)
            partSplit = fileName.split(".")
            df.to_excel(f"output\\{partSplit[0]}.xlsx",sheet_name=f"{partSplit[0]}", index=False)
        print("Convert all file json to excel!")
    else:
        print("Không có file json hợp lệ!")



def collect_ho_so_nhan_su():
    data_file = "output\\HO_SO_NHAN_SU.xlsx"
    data_ho_so_nhan_su = pd.read_excel(data_file)
    data_ho_so_nhan_su = data_ho_so_nhan_su[["id", "properties.Mã nhân viên.unique_id.prefix", "properties.Mã nhân viên.unique_id.number",
                                             "properties.Họ và tên.title", "properties.Quê quán.rich_text",                                          
                                             "properties.SĐT.phone_number", "properties.Email.email", 
                                             "properties.Ngày sinh.date.start", "properties.Chức vụ.select.name", 
                                             "properties.Ngày bắt đầu làm việc.date.start", "properties.Cơ sở.select.name", 
                                             "properties.Phân cấp.select.name"]]
    data_ho_so_nhan_su = data_ho_so_nhan_su.rename(columns={"id":"notion id", "properties.Mã nhân viên.unique_id.prefix":"Tiền tố", "properties.Mã nhân viên.unique_id.number":"Mã nhân viên",
                                             "properties.Họ và tên.title":"Họ và tên", "properties.Quê quán.rich_text":"Quê quán",                                          
                                             "properties.SĐT.phone_number":"SĐT", "properties.Email.email":"Email", 
                                             "properties.Ngày sinh.date.start":"Ngày sinh", "properties.Chức vụ.select.name":"Chức vụ", 
                                             "properties.Ngày bắt đầu làm việc.date.start":"Ngày bắt đầu làm việc", "properties.Cơ sở.select.name":"Cơ sở", 
                                             "properties.Phân cấp.select.name":"Phân cấp"})

    data_ho_so_nhan_su['Họ và tên'] = data_ho_so_nhan_su['Họ và tên'].apply(extract_text_content)
    data_ho_so_nhan_su["Quê quán"] = data_ho_so_nhan_su["Quê quán"].apply(extract_text_content)
    # Xử lý cột SĐT
    def extract_SDT(item):
        if(item is None or np.isnan(item)):
            return ""
        else:
            return f"0{round(item)}"
    data_ho_so_nhan_su["SĐT"] = data_ho_so_nhan_su["SĐT"].apply(extract_SDT).astype(str)
    return data_ho_so_nhan_su.sort_values("Mã nhân viên")

def collect_thong_tin_khach_hang():
    dataFile = "output\\THONG_TIN_KHACH_HANG.xlsx"
    dataThongTinKhachHang = pd.read_excel(dataFile)
    dataThongTinKhachHang = dataThongTinKhachHang[["id", "properties.Mã khách hàng.unique_id.prefix", 
                                           "properties.Mã khách hàng.unique_id.number", "properties.Họ và tên.title", 
                                           "properties.Cơ sở.rollup.array", "properties.CCCD.rich_text",  "properties.SĐT.phone_number", 
                                           "properties.Link Facebook.url", "properties.Địa chỉ.rich_text"
                                           ]]     
    dataThongTinKhachHang = dataThongTinKhachHang.rename(columns={"id":"notion id", "properties.Mã khách hàng.unique_id.prefix":"Tiền tố", 
                                           "properties.Mã khách hàng.unique_id.number":"Mã khách hàng", "properties.Họ và tên.title":"Họ và tên", 
                                           "properties.CCCD.rich_text":"CCCD",  "properties.SĐT.phone_number":"SĐT", 
                                           "properties.Link Facebook.url":"Facebook", "properties.Địa chỉ.rich_text":"Địa chỉ", "properties.Cơ sở.rollup.array" : "Cơ sở"})

    dataThongTinKhachHang["Họ và tên"] = dataThongTinKhachHang["Họ và tên"].apply(extract_plain_text)
    dataThongTinKhachHang["CCCD"] = dataThongTinKhachHang["CCCD"].apply(extract_text_content)
    dataThongTinKhachHang["Địa chỉ"] = dataThongTinKhachHang["Địa chỉ"].apply(extract_text_content)
    dataThongTinKhachHang["Cơ sở"] = dataThongTinKhachHang["Cơ sở"].apply(extract_select_name)
    return dataThongTinKhachHang

def collect_danh_muc_dich_vu():
    dataFile = "output\\DANH_MUC_DICH_VU.xlsx"
    dataDanhMucDichVu = pd.read_excel(dataFile)
    dataDanhMucDichVu = dataDanhMucDichVu[["id", "properties.Tên dịch vụ.title", "properties.Nhóm dịch vụ.select.name",
                                           "properties.Số ca.rollup.number", "properties.Công phụ phẫu 1.number",
                                           "properties.Công phụ phẫu 2.number"]]
                                           
    dataDanhMucDichVu = dataDanhMucDichVu.rename(columns={"id":"notion id", "properties.Tên dịch vụ.title": "Tên dịch vụ", 
                                                          "properties.Nhóm dịch vụ.select.name":"Nhóm dịch vụ",
                                                            "properties.Số ca.rollup.number":"Số ca", 
                                                            "properties.Công phụ phẫu 1.number":"Công phụ phẫu 1",
                                                            "properties.Công phụ phẫu 2.number":"Công phụ phẫu 2"})
    dataDanhMucDichVu['Tên dịch vụ'] = dataDanhMucDichVu['Tên dịch vụ'].apply(extract_text_content)

    return dataDanhMucDichVu.sort_values("Tên dịch vụ")

def collect_chi_tieu():
    dataFile = "output\\CHI_TIEU.xlsx"
    dataChiTieu = pd.read_excel(dataFile)
    dataChiTieu = dataChiTieu[["id", "properties.Auto mã chi tiêu.unique_id.prefix", "properties.Auto mã chi tiêu.unique_id.number",
                               "properties.Ngày chi.date.start", "properties.Cơ sở.select.name","properties.Phân loại.select.name", "properties.Nhân viên xác nhận.relation",
                               "properties.Lượng chi.number", "properties.Người Nhận/Ứng.relation"]]
    dataChiTieu = dataChiTieu.rename(columns={"id":"notion id", "properties.Ngày chi.date.start":"Ngày chi", "properties.Nhân viên xác nhận.relation":"id nhân viên xác nhận",
                               "properties.Cơ sở.select.name":"Cơ sở", "properties.Người Nhận/Ứng.relation":"id người nhận/ứng", 
                               "properties.Phân loại.select.name":"Phân loại", "properties.Auto mã chi tiêu.unique_id.prefix":"Tiền tố", 
                               "properties.Auto mã chi tiêu.unique_id.number":"Mã chi tiêu", "properties.Lượng chi.number":"Lượng chi"})

    dataChiTieu["id nhân viên xác nhận"] = dataChiTieu["id nhân viên xác nhận"].apply(extract_id)
    dataChiTieu["id người nhận/ứng"] = dataChiTieu["id người nhận/ứng"].apply(extract_id)
    dataChiTieu = pd.merge(dataChiTieu, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id nhân viên xác nhận", right_on="notion id", how="outer")
    dataChiTieu = dataChiTieu.rename(columns={"Họ và tên":"Nhân viên xác nhận"})
    dataChiTieu = pd.merge(dataChiTieu, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id người nhận/ứng", right_on="notion id", how="outer")
    dataChiTieu = dataChiTieu.rename(columns={"Họ và tên":"Người nhận/ứng"})
    dataChiTieu = dataChiTieu.drop(columns=["notion id"])

    dataChiTieu = dataChiTieu.drop(columns=["notion id_y"])
    dataChiTieu = dataChiTieu.rename(columns={"notion id_x":"notion id"})
    return dataChiTieu.sort_values("Mã chi tiêu")

def collect_doanh_thu_he_thong():
    dataFile = "output\\DOANH_THU_HE_THONG.xlsx"
    dataDoanhThuHeThong = pd.read_excel(dataFile)
    dataDoanhThuHeThong = dataDoanhThuHeThong[["id", "properties.Auto mã dịch vụ.unique_id.prefix", 
                                               "properties.Auto mã dịch vụ.unique_id.number", "properties.Ngày thực hiện.date.start", 
                                               "properties.Cơ sở.select.name", "properties.Loại hình dịch vụ.relation", "properties.Khách hàng.relation", "properties.Nguồn khách.select.name",
                                               "properties.Sale chính.relation", "properties.Đơn giá gốc.number",
                                               "properties.Sale phụ.relation","properties.Upsale.number","properties.Đơn giá.formula.number",
                                                "properties.Thanh toán lần đầu.number", "properties.Trả sau.rollup.number",
                                                "properties.Đã thanh toán.formula.number", "properties.Dư nợ.formula.number",
                                               "properties.Bác sĩ 1.relation",  "properties.Bác sĩ 2.relation",
                                                "properties.Phụ phẫu 1.relation", "properties.Phụ phẫu 2.relation"
                                               ]]
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"id":"notion id", "properties.Phụ phẫu 1.relation":"id phụ phẫu 1", 
                                               "properties.Khách hàng.relation":"id khách hàng", "properties.Auto mã dịch vụ.unique_id.prefix":"Tiền tố", 
                                               "properties.Auto mã dịch vụ.unique_id.number":"Mã dịch vụ", "properties.Đã thanh toán.formula.number":"Đã thanh toán", 
                                               "properties.Cơ sở.select.name":"Cơ sở", "properties.Bác sĩ 1.relation":"id bác sĩ 1", 
                                               "properties.Thanh toán lần đầu.number":"Thanh toán lần đầu", "properties.Đơn giá.formula.number":"Đơn giá", 
                                               "properties.Dư nợ.formula.number":"Dư nợ", "properties.Phụ phẫu 2.relation":"id phụ phẫu 2", 
                                               "properties.Bác sĩ 2.relation":"id bác sĩ 2", "properties.Sale chính.relation":"id sale chính", "properties.Đơn giá gốc.number":"Đơn giá gốc",
                                               "properties.Sale phụ.relation":"id sale phụ","properties.Upsale.number":"Upsale",
                                               "properties.Nguồn khách.select.name":"Nguồn khách", "properties.Trả sau.rollup.number":"Trả sau", 
                                               "properties.Ngày thực hiện.date.start":"Ngày thực hiện", "properties.Loại hình dịch vụ.relation":"id loại hình dịch vụ"})
        # Xử lý lấy id nhân sự
    dataDoanhThuHeThong = dataDoanhThuHeThong.fillna("")

    dataDoanhThuHeThong["id khách hàng"] = dataDoanhThuHeThong["id khách hàng"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_thong_tin_khach_hang()[["notion id", "Họ và tên"]], left_on="id khách hàng", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"Họ và tên":"Khách hàng"})
    dataDoanhThuHeThong["id sale chính"] = dataDoanhThuHeThong["id sale chính"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id sale chính", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"Họ và tên":"Sale chính"})
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id"])
    dataDoanhThuHeThong["id sale phụ"] = dataDoanhThuHeThong["id sale phụ"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id sale phụ", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"Họ và tên":"Sale phụ"})
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id"])
    dataDoanhThuHeThong["id bác sĩ 1"] = dataDoanhThuHeThong["id bác sĩ 1"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id bác sĩ 1", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"Họ và tên":"Bác sĩ 1"})
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id"])
    dataDoanhThuHeThong["id bác sĩ 2"] = dataDoanhThuHeThong["id bác sĩ 2"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id bác sĩ 2", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"Họ và tên":"Bác sĩ 2"})
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id"])
    dataDoanhThuHeThong["id phụ phẫu 1"] = dataDoanhThuHeThong["id phụ phẫu 1"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id phụ phẫu 1", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"Họ và tên":"Phụ phẫu 1"})
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id"])
    dataDoanhThuHeThong["id phụ phẫu 2"] = dataDoanhThuHeThong["id phụ phẫu 2"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id phụ phẫu 2", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"Họ và tên":"Phụ phẫu 2"})
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id"])
    dataDoanhThuHeThong["id loại hình dịch vụ"] = dataDoanhThuHeThong["id loại hình dịch vụ"].apply(extract_id)
    dataDoanhThuHeThong = pd.merge(dataDoanhThuHeThong, collect_danh_muc_dich_vu()[["notion id", "Tên dịch vụ"]], left_on="id loại hình dịch vụ", right_on="notion id", how="outer")
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id"])
    
    dataDoanhThuHeThong = dataDoanhThuHeThong.drop(columns=["notion id_y"])
    dataDoanhThuHeThong = dataDoanhThuHeThong.rename(columns={"notion id_x":"notion id"})
    return dataDoanhThuHeThong.sort_values("Mã dịch vụ")


def collect_danh_sach_thu_no():
    dataFile = "output\\DANH_SACH_THU_NO.xlsx"
    dataThuNo = pd.read_excel(dataFile)
    dataThuNo = dataThuNo[["id", "properties.Gen mã đơn.unique_id.prefix", "properties.Gen mã đơn.unique_id.number",
                           "properties.Ngày thu.date.start", "properties.Cơ sở.rollup.array",
                           "properties.Đơn nợ.relation", "properties.Lượng thu.number",
                           "properties.Người thu.relation",
                           "properties.Sale.rollup.array"]]
    dataThuNo = dataThuNo.rename(columns={"id":"notion id", "properties.Gen mã đơn.unique_id.prefix":"Tiền tố", "properties.Gen mã đơn.unique_id.number": "Mã đơn thu nợ",
                           "properties.Ngày thu.date.start":"Ngày thu", "properties.Cơ sở.rollup.array":"Cơ sở",
                           "properties.Đơn nợ.relation":"id đơn nợ",
                           "properties.Người thu.relation":"id người thu", "properties.Lượng thu.number":"Lượng thu",
                           "properties.Sale.rollup.array":"id sale"})
    def extract_relation_id(item):
        item = convert_data_in_cell_to_json(item)
        if len(item) > 0:
            if len(item[0]["relation"]):
                return item[0]["relation"][0]["id"]
            else:
                 return ""
        else:
            return 0

    dataThuNo["id đơn nợ"] = dataThuNo["id đơn nợ"].apply(extract_id)
    dataThuNo["id người thu"] = dataThuNo["id người thu"].apply(extract_id)
    dataThuNo["Cơ sở"] = dataThuNo["Cơ sở"].apply(extract_select_name)
    dataThuNo["id sale"] = dataThuNo["id sale"].apply(extract_relation_id)

    dataThuNo = pd.merge(dataThuNo, collect_ho_so_nhan_su()[["notion id", "Họ và tên"]], left_on="id sale", right_on="notion id", how="outer")
    dataThuNo = dataThuNo.rename(columns={"Họ và tên":"Sale"})
    dataThuNo = pd.merge(dataThuNo, collect_doanh_thu_he_thong()[["notion id", "Mã dịch vụ"]], left_on="id đơn nợ", right_on="notion id", how="outer")
    dataThuNo = dataThuNo.rename(columns={"Mã dịch vụ":"Đơn nợ"})
    def format_don_no(item):
        item = '{:.0f}'.format(item)
        return f"HD-LUXURY-{item}"
    dataThuNo["Đơn nợ"] = dataThuNo["Đơn nợ"].apply(format_don_no)

    
    dataThuNo = dataThuNo.drop(columns=["notion id_y"])
    dataThuNo = dataThuNo.drop(columns=["notion id"])
    dataThuNo = dataThuNo.rename(columns={"notion id_x":"notion id"})
    
    return dataThuNo.sort_values("Mã đơn thu nợ")

def collect_data():
    convert_json_to_excel()
    # Kiểm tra xem file Excel đã tồn tại hay chưa
    excelFilePath = "output\\ALL.xlsx"
    if os.path.exists(excelFilePath):
        # Nếu đã tồn tại, xóa file cũ đi
        try:
            os.remove(excelFilePath)
            print(f"Đã xóa file Excel cũ '{excelFilePath}'")
        except Exception as e:
            print(f"Lỗi khi xóa file Excel cũ: {e}")

    # Tạo workbook mới
    wb = Workbook()
    # Tạo sheet hồ sơ nhân sự
    ws1 = wb.active
    ws1.title = 'Hồ sơ nhân sự'
    writeDataframeToSheet(ws1, collect_ho_so_nhan_su())
    # Tạo sheet danh sách khách hàng
    ws2 = wb.create_sheet("Danh sách khách hàng")
    writeDataframeToSheet(ws2, collect_thong_tin_khach_hang())
    # Tạo sheet danh mục dịch vụ
    ws3 = wb.create_sheet("Danh mục dịch vụ")
    writeDataframeToSheet(ws3, collect_danh_muc_dich_vu())
    # Tạo sheet chi tiêu
    ws4 = wb.create_sheet(title='Chi tiêu')
    writeDataframeToSheet(ws4, collect_chi_tieu())
    # Tạo sheet doanh thu hệ thống
    ws5 = wb.create_sheet(title="Doanh thu hệ thống")
    writeDataframeToSheet(ws5, collect_doanh_thu_he_thong())
    # Tạo sheet Thu nợ
    ws6 = wb.create_sheet(title="Thu nợ")
    writeDataframeToSheet(ws6, collect_danh_sach_thu_no())



    # Lưu workbook vào file Excel
    try:
        wb.save(excelFilePath)
        print(f"Đã tạo file Excel mới '{excelFilePath}' thành công")
    except Exception as e:
        print(f"Lỗi khi tạo file Excel mới: {e}")


# collect_thong_tin_khach_hang()
# collect_doanh_thu_he_thong()
# collect_data()
    