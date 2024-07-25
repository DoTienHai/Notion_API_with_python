import os
from openpyxl import Workbook
import pandas as pd
from datetime import datetime
from Config import *

month = datetime.today().month
year = datetime.today().year

columns = ["Tiền tố", "Mã dịch vụ", "Ngày thực hiện",
                "Cơ sở", "Khách hàng", "Nguồn khách", "Nhóm dịch vụ", "Tên dịch vụ", "Sale chính", "Đơn giá gốc", 
                "Sale phụ", "Upsale", "Đơn giá", "Thanh toán lần đầu", "Trả sau",
                "Đã thanh toán", "Dư nợ", "Bác sĩ 1", "Bác sĩ 2", "Phụ phẫu 1", 
                "Phụ phẫu 2", "Công phụ phẫu 1", "Công phụ phẫu 2", 
                "Tỉ lệ chiết khấu sale chính", "Tỉ lệ chiết khấu sale phụ",
                "Chiết khấu sale chính", "Chiết khấu sale phụ", 
                "Tỉ lệ chiết khấu bác sĩ 1", "Tỉ lệ chiết khấu bác sĩ 2", 
                "Chiết khấu bác sĩ 1", "Chiết khấu bác sĩ 2"]

def filter_date(data, column_name):
    # column_name is name of column datetime in dataframe
    # convert data to datetime type
    data[column_name] = pd.to_datetime(data[column_name])
    # filter data in expected month 
    data = data[(data[column_name].dt.year == year) & (data[column_name].dt.month == month)]
    # Put the column containing the formatted datetime at the top
    data = data.rename(columns={column_name:f"{column_name}_temp"})
    data[column_name] = data[f"{column_name}_temp"].dt.strftime('%m-%d-%Y')
    data = data.drop(columns=[f"{column_name}_temp"])
    columns = [column_name] + [col for col in data.columns if col != column_name]
    data = data[columns]
    return data

def filter_date_don_no(data, column_name):
    # column_name is name of column datetime in dataframe
    # convert data to datetime type
    data[column_name] = pd.to_datetime(data[column_name])
    # filter data in expected month 
    data = data[(data[column_name].dt.year != year) | (data[column_name].dt.month != month)]
    # Put the column containing the formatted datetime at the top
    data = data.rename(columns={column_name:f"{column_name}_temp"})
    data[column_name] = data[f"{column_name}_temp"].dt.strftime('%m-%d-%Y')
    data = data.drop(columns=[f"{column_name}_temp"])
    columns = [column_name] + [col for col in data.columns if col != column_name]
    data = data[columns]
    return data

def add_total_row(data):
    sum_data = data.select_dtypes(include=['number']).sum()
    count_col = ""
    for col in data.columns.tolist():
        if "Mã" in col:
            count_col = col
            break
    sum_data[count_col] = data[count_col].count()  
    total_df = pd.DataFrame(sum_data).T
    # Thêm các cột không phải là số vào dòng tổng
    for col in data.columns:
        if col not in total_df.columns:
            total_df[col] = ''  
    # Đặt lại thứ tự các cột để khớp với DataFrame gốc
    total_df = total_df[data.columns]
    total_df["Tiền tố"] = "Tổng"
    # Nối dòng tổng với DataFrame gốc
    data = pd.concat([data, total_df])
    return data 


def get_don_sale_chinh(notion_id_nhan_su):
    data = get_data_doanh_thu("",["ALL"])
    data = data[data["id sale chính"] == notion_id_nhan_su]
    data = filter_date(data, "Ngày thực hiện")
    data = data[columns]
    data = add_total_row(data)
    
    return data

def get_don_sale_phu(notion_id_nhan_su):
    data = get_data_doanh_thu("",["ALL"])
    data = data[data["id sale phụ"] == notion_id_nhan_su]
    data = filter_date(data, "Ngày thực hiện")
    data = data[columns]
    data = add_total_row(data) 
    return data

def get_don_1_bac_si(notion_id_nhan_su):
    data = get_data_doanh_thu("",["ALL"])
    data = data[(data["id bác sĩ 1"] == notion_id_nhan_su) & (data["id bác sĩ 2"].isnull())]
    data = filter_date(data, "Ngày thực hiện")
    data = data[columns]
    data = add_total_row(data) 
    return data

def get_don_2_bac_si(notion_id_nhan_su):
    data = get_data_doanh_thu("",["ALL"])
    data = data[(data["id bác sĩ 1"] == notion_id_nhan_su)& ~(data["id bác sĩ 2"].isnull()) | (data["id bác sĩ 2"] == notion_id_nhan_su)]
    data = filter_date(data, "Ngày thực hiện")
    data = data[columns]
    data = add_total_row(data) 
    return data

def get_don_phụ_phau_1(notion_id_nhan_su):
    data = get_data_doanh_thu("",["ALL"])
    data = data[data["id phụ phẫu 1"] == notion_id_nhan_su]
    data = filter_date(data, "Ngày thực hiện")
    data = data[columns]
    data = add_total_row(data) 
    return data

def get_don_phụ_phau_2(notion_id_nhan_su):
    data = get_data_doanh_thu("",["ALL"])
    data = data[data["id phụ phẫu 2"] == notion_id_nhan_su]
    data = filter_date(data, "Ngày thực hiện")
    data = data[columns]
    data = add_total_row(data) 

    return data

def get_don_thu_no(notion_id_nhan_su):
    data = get_data_thu_no("", ["ALL"])
    data = filter_date(data, "Ngày thu")
    data = filter_date_don_no(data, "Ngày thực hiện")
    data = data[(data["id sale chính"] == notion_id_nhan_su) | (data["id sale phụ"] == notion_id_nhan_su) | (data["id bác sĩ 1"] == notion_id_nhan_su) | (data["id bác sĩ 2"] == notion_id_nhan_su)]

    if len(data):
        for index, row in data.iterrows():
            if row["id sale chính"] != notion_id_nhan_su:
                data.at[index, "Tỉ lệ chiết khấu sale chính"] = 0
                data.at[index, "Chiết khấu sale chính"] = 0
            if row["id sale phụ"] != notion_id_nhan_su:
                data.at[index, "Tỉ lệ chiết khấu sale phụ"] = 0
                data.at[index, "Chiết khấu sale phụ"] = 0
            if row["id bác sĩ 1"] != notion_id_nhan_su:
                data.at[index, "Tỉ lệ chiết khấu bác sĩ 1"] = 0
                data.at[index, "Chiết khấu bác sĩ 1"] = 0
            if row["id bác sĩ 2"] != notion_id_nhan_su:
                data.at[index, "Tỉ lệ chiết khấu bác sĩ 2"] = 0
                data.at[index, "Chiết khấu bác sĩ 2"] = 0

        sum_data = data.select_dtypes(include=['number']).sum()
        sum_data["Mã đơn thu nợ"] = data["Mã đơn thu nợ"].count()  
        total_df = pd.DataFrame(sum_data).T
        # Thêm các cột không phải là số vào dòng tổng
        for col in data.columns:
            if col not in total_df.columns:
                total_df[col] = ''  
        # Đặt lại thứ tự các cột để khớp với DataFrame gốc
        total_df = total_df[data.columns]
        total_df["Tiền tố"] = "Tổng"
        # Nối dòng tổng với DataFrame gốc
        data = pd.concat([data, total_df])
    return data
    

def create_tong_hop_luong_co_so(path, location):
    path = os.path.join(path, "Báo cáo cá nhân")
    list_of_report_ca_nhan_path = []
    for root, dir, files in os.walk(path):
        for file in files:
            if file.endswith(".xlsx") and ("Tổng hợp lương nhân viên" not in file):
                list_of_report_ca_nhan_path.append(os.path.join(root, file))

    cols = ["Mã nhân viên", "Tên nhân viên", f"Tổng lương tại {location}"]
    data = pd.DataFrame()
    for report_ca_nhan_path in list_of_report_ca_nhan_path:
        file_name = os.path.basename(report_ca_nhan_path)
        file_name_part = file_name.split(" ")
        ma_nhan_vien = file_name_part[0]
        ten_nhan_vien = ' '.join(file_name_part[1:-1])
        data_luong = pd.read_excel(report_ca_nhan_path, sheet_name="Lương")
        tong_luong_tai_co_so = data_luong.set_index("Danh mục lương").transpose()[f"Tổng lương tại {location}"]
        if len(tong_luong_tai_co_so):
            tong_luong_tai_co_so = tong_luong_tai_co_so.values[0]
        else:
            tong_luong_tai_co_so = 0
        # print(tong_luong_tai_co_so)
        row_data = {
            "Mã nhân viên" : [ma_nhan_vien], 
            "Tên nhân viên" : [ten_nhan_vien], 
            f"Tổng lương tại {location}" : [tong_luong_tai_co_so]
        }
        df_row_data = pd.DataFrame(row_data, columns=cols)
        data = pd.concat([data, df_row_data])
    
    Tong_luong = data[f"Tổng lương tại {location}"].sum()
    row_data = {
        "Mã nhân viên" : "Tổng lương", 
        "Tên nhân viên" : [""], 
        f"Tổng lương tại {location}" : [Tong_luong]
    }
    df_row_data = pd.DataFrame(row_data, columns=cols)
    data = pd.concat([data, df_row_data])
        
    # Kiểm tra xem file Excel đã tồn tại hay chưa
    excel_file_path = os.path.join(path, f"Tổng hợp lương nhân viên tại {location} {file_name_part[-1]}.xlsx")
    if os.path.exists(excel_file_path):
        # Nếu đã tồn tại, xóa file cũ đi
        try:
            os.remove(excel_file_path)
            print(f"Đã xóa file Excel cũ '{excel_file_path}'")
        except Exception as e:
            print(f"Lỗi khi xóa file Excel cũ: {e}")
    # Tạo workbook mới
    wb = Workbook()
    # Tạo sheet Đơn sale chính
    ws1 = wb.active
    ws1.title = 'Tổng hợp lương'
    if len(data) > 1:
        writeDataframeToSheet(ws1, data)

    # Lưu workbook vào file Excel
    try:
        wb.save(excel_file_path)
        print(f"Đã tạo file Excel mới '{excel_file_path}' thành công")
    except Exception as e:
        print(f"Lỗi khi tạo file Excel mới: {e}")
 

def create_report_ca_nhan(path, info_nhan_su):
    notion_id_nhan_su = info_nhan_su["notion id"]
    ho_va_ten = info_nhan_su["Họ và tên"]
    ma_nhan_vien = f"{info_nhan_su["Tiền tố"]}-{info_nhan_su["Mã nhân viên"]}"
    co_so = info_nhan_su["Cơ sở"]
    # Kiểm tra xem file Excel đã tồn tại hay chưa
    folder_path = os.path.join(path,"Báo cáo cá nhân")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    excel_file_path = os.path.join(folder_path,f"{ma_nhan_vien} {ho_va_ten} {month}-{year}.xlsx")
    if os.path.exists(excel_file_path):
        # Nếu đã tồn tại, xóa file cũ đi
        try:
            os.remove(excel_file_path)
            print(f"Đã xóa file Excel cũ '{excel_file_path}'")
        except Exception as e:
            print(f"Lỗi khi xóa file Excel cũ: {e}")

    # Tạo workbook mới
    wb = Workbook()
    # Tạo sheet Đơn sale chính
    ws1 = wb.active
    ws1.title = 'Đơn sale chính'
    data_sale_chinh = get_don_sale_chinh(notion_id_nhan_su)
    if len(data_sale_chinh) > 1:
        writeDataframeToSheet(ws1, data_sale_chinh[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở", "Khách hàng", "Nguồn khách", 
                                                    "Tên dịch vụ", "Đơn giá gốc", "Sale phụ", "Upsale", "Đơn giá", "Đã thanh toán",
                                                    "Tỉ lệ chiết khấu sale chính", "Chiết khấu sale chính"]])
    # Tạo sheet Đơn sale phụ
    data_sale_phu = get_don_sale_phu(notion_id_nhan_su)
    if len(data_sale_phu) > 1:
        ws2 = wb.create_sheet(title='Đơn sale phụ')
        writeDataframeToSheet(ws2, data_sale_phu[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở", "Khách hàng", "Nguồn khách", 
                                                    "Tên dịch vụ", "Đơn giá gốc", "Sale phụ", "Upsale", "Đơn giá", "Đã thanh toán",
                                                    "Tỉ lệ chiết khấu sale phụ", "Chiết khấu sale phụ"]])
    # Tạo sheet Đơn 1 bác sĩ 
    data_don_1_bac_si = get_don_1_bac_si(notion_id_nhan_su)
    if len(data_don_1_bac_si) > 1:
        ws3 = wb.create_sheet(title="Đơn 1 bác sĩ")
        writeDataframeToSheet(ws3, data_don_1_bac_si[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở", "Khách hàng", "Nguồn khách", 
                                                    "Tên dịch vụ", "Đơn giá gốc", "Sale phụ", "Upsale", "Đơn giá", "Đã thanh toán",
                                                    "Tỉ lệ chiết khấu bác sĩ 1", "Chiết khấu bác sĩ 1"]])
    # Tạo sheet Đơn 2 bác sĩ
    data_don_2_bac_si = get_don_2_bac_si(notion_id_nhan_su)
    if len(data_don_2_bac_si) > 1:
        ws4 = wb.create_sheet(title="Đơn 2 bác sĩ")
        writeDataframeToSheet(ws4, data_don_2_bac_si[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở", "Khách hàng", "Nguồn khách", 
                                                    "Tên dịch vụ", "Đơn giá gốc", "Sale phụ", "Upsale", "Đơn giá", "Đã thanh toán",
                                                    "Tỉ lệ chiết khấu bác sĩ 2", "Chiết khấu bác sĩ 2"]])
    # Tạo sheet Đơn phụ phẫu 1
    data_phu_phau_1 = get_don_phụ_phau_1(notion_id_nhan_su)
    if len(data_phu_phau_1) > 1:
        ws5 = wb.create_sheet("Đơn phụ phẫu 1")
        writeDataframeToSheet(ws5, data_phu_phau_1[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở", "Khách hàng", "Nguồn khách", 
                                                    "Tên dịch vụ", "Phụ phẫu 1", "Công phụ phẫu 1"]])
    # Tạo sheet Đơn phụ phẫu 2
    data_phu_phau_2 = get_don_phụ_phau_2(notion_id_nhan_su)
    if len(data_phu_phau_2) > 1:
        ws6 = wb.create_sheet("Đơn phụ phẫu 2")
        writeDataframeToSheet(ws6, data_phu_phau_2[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở", "Khách hàng", "Nguồn khách", 
                                                    "Tên dịch vụ", "Phụ phẫu 2", "Công phụ phẫu 2"]])
    # Tạo sheet Đơn thu nợ
    data_don_thu_no = get_don_thu_no(notion_id_nhan_su)
    if len(data_don_thu_no) > 1:
        ws7 = wb.create_sheet("Đơn thu nợ")
        writeDataframeToSheet(ws7, data_don_thu_no[["Tiền tố", "Mã đơn thu nợ", "Lượng thu", "Ngày thu", "Cơ sở", "Đơn nợ", "Tên dịch vụ", "Khách hàng", "Nguồn khách", 
                                                    "Sale chính", "Đơn giá gốc", "Sale phụ", "Upsale", "Đơn giá", "Đã thanh toán", "Bác sĩ 1", "Bác sĩ 2",
                                                    "Tỉ lệ chiết khấu sale chính", "Chiết khấu sale chính", 
                                                    "Tỉ lệ chiết khấu sale phụ", "Chiết khấu sale phụ", 
                                                    "Tỉ lệ chiết khấu bác sĩ 1", "Chiết khấu bác sĩ 1",
                                                    "Tỉ lệ chiết khấu bác sĩ 2", "Chiết khấu bác sĩ 2"]])
    # Tạo sheet tính lương
    if co_so != "OUTSIDE":
        ref_luong = pd.read_excel("Ref tính lương.xlsx", sheet_name="Lương cơ bản")
        ws8 = wb.create_sheet("Lương")
        data_luong = pd.DataFrame()
        # Tính lương cơ bản theo ngày công
        if info_nhan_su["Hình thức làm việc"] == "Remote":
            # Lấy số ngày lương
            data_luong["Ngày công"] = 28
            data_luong["Phụ cấp"] = 0
        else:
            # Lấy số ngày lương
            data_cham_cong = get_data_cham_cong_tong_hop()
            data_luong["Ngày công"] = data_cham_cong[data_cham_cong["id nhân sự"] == notion_id_nhan_su]["Tổng công"]
            data_luong["Phụ cấp"] = data_luong["Ngày công"]*35000
        
        luong_co_ban = ref_luong[ref_luong["notion id"] == notion_id_nhan_su]["Tổng lương cơ bản"]
        if len(luong_co_ban):
            luong_co_ban = luong_co_ban.values[0]
        ngay_cong = data_luong["Ngày công"]
        if len(ngay_cong):
            ngay_cong = ngay_cong.values[0]
        tong_luong_co_ban = luong_co_ban*ngay_cong/28
        for location in location_list:
            if location != "HỆ THỐNG":
                ti_le_luong = ref_luong[ref_luong["notion id"] == notion_id_nhan_su][location]
                if len(ti_le_luong):
                    ti_le_luong = ti_le_luong.values[0]
                data_luong[f"Lương cơ bản tại {location}"] = tong_luong_co_ban*ti_le_luong
        
        # Tính chiết khấu doanh số kinh doanh
            if len(data_sale_chinh):
                data_luong[f"Chiết khấu sale chính tại {location}"] = data_sale_chinh[data_sale_chinh["Cơ sở"] == location]["Chiết khấu sale chính"].sum()
            else:
                data_luong[f"Chiết khấu sale chính tại {location}"] = 0
            if len(data_sale_phu):
                data_luong[f"Chiết khấu sale phụ tại {location}"] = data_sale_phu[data_sale_phu["Cơ sở"] == location]["Chiết khấu sale phụ"].sum()
            else:
                data_luong[f"Chiết khấu sale phụ tại {location}"] = 0
        # Tính chiết khấu phẫu thuật
            if len(data_don_1_bac_si):
                data_luong[f"Đơn 1 bác sĩ tại {location}"] = data_don_1_bac_si[data_don_1_bac_si["Cơ sở"] == location]["Chiết khấu bác sĩ 1"].sum()
            else:
                data_luong[f"Đơn 1 bác sĩ tại {location}"] = 0

            if len(data_don_2_bac_si):
                data_luong[f"Đơn 2 bác sĩ tại {location}"] = data_don_2_bac_si[data_don_2_bac_si["Cơ sở"] == location]["Chiết khấu bác sĩ 2"].sum()
            else:
                data_luong[f"Đơn 2 bác sĩ tại {location}"] = 0       
        # Tính công phụ phẫu
            if len(data_phu_phau_1):
                data_luong[f"Công phụ phẫu 1 tại {location}"] = data_phu_phau_1[data_phu_phau_1["Cơ sở"] == location]["Công phụ phẫu 1"].sum()
            else:
                data_luong[f"Công phụ phẫu 1 tại {location}"] = 0

            if len(data_phu_phau_2):
                data_luong[f"Công phụ phẫu 2 tại {location}"] = data_phu_phau_2[data_phu_phau_2["Cơ sở"] == location]["Công phụ phẫu 2"].sum()
            else:
                data_luong[f"Công phụ phẫu 2 tại {location}"] = 0   
        # Tính chiết khấu thu nợ
            if len(data_don_thu_no):
                data_luong[f"Chiết khấu thu nợ tại {location}"] = data_don_thu_no[(data_don_thu_no["Cơ sở"] == location)]["Chiết khấu sale chính"].sum() + \
                                                                    data_don_thu_no[(data_don_thu_no["Cơ sở"] == location)]["Chiết khấu sale phụ"].sum() + \
                                                                    data_don_thu_no[(data_don_thu_no["Cơ sở"] == location)]["Chiết khấu bác sĩ 1"].sum() + \
                                                                    data_don_thu_no[(data_don_thu_no["Cơ sở"] == location)]["Chiết khấu bác sĩ 2"].sum()
                                                                    
        # Ứng lương
            data_ung_luong = get_data_chi_tieu("", ["ALL"])
            data_ung_luong = filter_date(data_ung_luong, "Ngày chi")
            data_ung_luong = data_ung_luong[(data_ung_luong["Phân loại"] == "Ứng Lương") & (data_ung_luong["id người nhận/ứng"] == notion_id_nhan_su)]
            if len(data_ung_luong):
                data_luong[f"Ứng lương tại {location}"] = -data_ung_luong[data_ung_luong["Cơ sở"] == location]["Lượng chi"].sum()
            else:
                data_luong[f"Ứng lương tại {location}"] = 0
        # Thưởng
        # Phạt
        # khác 

        # Tổng kết lương theo cơ sở
        for location in location_list:
            data_luong[f"Tổng lương tại {location}"] = 0
            for col in data_luong.columns.tolist():
                if location in col:
                    data_luong[f"Tổng lương tại {location}"] += data_luong[col].sum()
            if location == co_so:
                data_luong[f"Tổng lương tại {location}"] += data_luong["Phụ cấp"]*2
        # Tổng lương
        data_luong["Tổng lương"] = 0
        for location in location_list:
            data_luong[f"Tổng lương tại {location}"] = data_luong[f"Tổng lương tại {location}"]/2
            data_luong["Tổng lương"] = data_luong["Tổng lương"] + data_luong[f"Tổng lương tại {location}"]
        data_luong_T = data_luong.transpose()
        # Đặt lại tên cột
        data_luong_T.columns = data_luong.index
        data_luong_T = data_luong_T.reset_index()
        data_luong_T.rename(columns={'index': 'Danh mục lương'}, inplace=True)
        writeDataframeToSheet(ws8, data_luong_T)
        # print(data_luong)


    # Lưu workbook vào file Excel
    try:
        wb.save(excel_file_path)
        print(f"Đã tạo file Excel mới '{excel_file_path}' thành công")
    except Exception as e:
        print(f"Lỗi khi tạo file Excel mới: {e}")

    # Tổng hợp lương của cơ sở
    # create_tong_hop_luong_co_so(folder_path, co_so)

# create_doanh_so_ca_nhan()
# create_tong_hop_luong_co_so()