import os
from openpyxl import Workbook
import pandas as pd
from Config import *
from UpdateLuyKe import get_data_cho_luy_ke
from CreateReportCaNhan import filter_date, add_total_row, filter_date_don_no

def get_data_chi_tiet_doanh_thu(location):
    data = get_data_doanh_thu(location, ["ALL"])
    data = data[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở","Tên dịch vụ", "Khách hàng",
                 "Nguồn khách", "Sale chính", "Đơn giá gốc", "Sale phụ", "Upsale", "Đơn giá", 
                 "Bác sĩ 1", "Bác sĩ 2", "Thanh toán lần đầu", "Trả sau", "Đã thanh toán", 
                 "Dư nợ", "Phụ phẫu 1", "Phụ phẫu 2", "Công phụ phẫu 1", "Công phụ phẫu 2"]]
    data = filter_date(data, "Ngày thực hiện")
    data = add_total_row(data)
    return data

def get_data_chi_tiet_thu_no(location):
    data = get_data_thu_no(location, ["ALL"])
    data = filter_date(data, "Ngày thu")
    data = data[["Tiền tố", "Mã đơn thu nợ", "Ngày thu", 
                 "Cơ sở", "Đơn nợ", "Ngày thực hiện","Lượng thu",
                 "Tên dịch vụ", "Khách hàng", "Nguồn khách", "Sale chính", 
                 "Đơn giá gốc", "Sale phụ", "Upsale", "Thanh toán lần đầu", 
                 "Đã thanh toán", "Bác sĩ 1", "Bác sĩ 2"]]
    data = add_total_row(data)
    return data

def get_data_chi_tiet_chi_tieu(location):
    data = get_data_chi_tieu(location, ["ALL"])
    data = filter_date(data, "Ngày chi")
    data = data[["Tiền tố", "Mã chi tiêu", "Ngày chi", "Cơ sở", "Phân loại", "Lượng chi", "Ghi chú"]]
    data = add_total_row(data)
    return data

def get_data_report_doanh_so(location):
    data = get_data_doanh_thu(location,["ALL"])
    if (location != "HỆ THỐNG"):
        data = data[data["Cơ sở"] == location]
    data = filter_date(data, "Ngày thực hiện")

    groupDataDoanhSo = pd.DataFrame(columns=["Nhân viên"])
    # Group data Sale chính
    groupDataSaleChinh = data.groupby("Sale chính")[["Đơn giá gốc"]].sum().reset_index()
    groupDataSaleChinh = groupDataSaleChinh.rename(columns={"Sale chính":"Nhân viên", "Đơn giá gốc":"Tổng đơn giá sale vòng 1"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataSaleChinh, on='Nhân viên', how='outer')

    # Group data Sale phụ
    groupDataSalePhu = data.groupby("Sale phụ")[["Upsale"]].sum().reset_index()
    groupDataSalePhu = groupDataSalePhu.rename(columns={"Sale phụ":"Nhân viên", "Upsale":"Tổng đơn giá vòng upsale"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataSalePhu, on='Nhân viên', how='outer')
    # Group data KPI

    # Group data 1 bác sĩ
    groupData1BacSi = data[data["id bác sĩ 2"].isnull()]
    groupData1BacSi = groupData1BacSi.groupby("Bác sĩ 1")[["Đã thanh toán"]].sum().reset_index()
    groupData1BacSi = groupData1BacSi.rename(columns={"Bác sĩ 1":"Nhân viên", "Đã thanh toán":"Doanh số đơn 1 bác sĩ"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupData1BacSi, on='Nhân viên', how='outer')
    # Group data 2 bác sĩ
    temp = data[~data["id bác sĩ 2"].isnull()]
    groupData2BacSi = temp.groupby("Bác sĩ 1")[["Đã thanh toán"]].sum().reset_index()
    groupData2BacSi = groupData2BacSi.rename(columns={"Bác sĩ 1":"Nhân viên", "Đã thanh toán":"A"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupData2BacSi, on='Nhân viên', how='outer')
    groupData2BacSi = temp.groupby("Bác sĩ 2")[["Đã thanh toán"]].sum().reset_index()
    groupData2BacSi = groupData2BacSi.rename(columns={"Bác sĩ 2":"Nhân viên", "Đã thanh toán":"B"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupData2BacSi, on='Nhân viên', how='outer')
    groupDataDoanhSo["Doanh số đơn 2 bác sĩ"] = groupDataDoanhSo["A"] + groupDataDoanhSo["B"]
    # group data phụ phẫu 1
    groupDataCountPhuPhau1 = data["Phụ phẫu 1"].value_counts().reset_index()
    groupDataCountPhuPhau1 = groupDataCountPhuPhau1.rename(columns={"Phụ phẫu 1":"Nhân viên", "count":"Số lần phụ phẫu 1"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCountPhuPhau1, on='Nhân viên', how='outer')
    # group data công phụ phẫu 1
    groupDataCongPhuPhau1 = data.groupby("Phụ phẫu 1")[["Công phụ phẫu 1"]].sum().reset_index()
    groupDataCongPhuPhau1 = groupDataCongPhuPhau1.rename(columns={"Phụ phẫu 1":"Nhân viên"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCongPhuPhau1, on='Nhân viên', how='outer')
    # group data phụ phẫu 2
    groupDataCountPhuPhau2 = data["Phụ phẫu 2"].value_counts().reset_index()
    groupDataCountPhuPhau2 = groupDataCountPhuPhau2.rename(columns={"Phụ phẫu 2":"Nhân viên", "count":"Số lần phụ phẫu 2"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCountPhuPhau2, on='Nhân viên', how='outer')
    # group data công phụ phẫu 2
    groupDataCongPhuPhau2 = data.groupby("Phụ phẫu 2")[["Công phụ phẫu 2"]].sum().reset_index()
    groupDataCongPhuPhau2 = groupDataCongPhuPhau2.rename(columns={"Phụ phẫu 2":"Nhân viên"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCongPhuPhau2, on='Nhân viên', how='outer')
    #group data thu nơ
    groupDataThuNo = get_data_thu_no(location, ["ALL"])
    groupDataThuNo = filter_date(groupDataThuNo, "Ngày thu")

    groupDataThuNo = groupDataThuNo.groupby("Sale chính")[["Lượng thu"]].sum().reset_index()
    groupDataThuNo = groupDataThuNo.rename(columns={"Sale chính":"Nhân viên", "Lượng thu":"Doanh số thu nợ"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataThuNo, on='Nhân viên', how='outer')
    groupDataDoanhSo = groupDataDoanhSo.drop(columns=["A","B"])
    groupDataDoanhSo = groupDataDoanhSo.fillna(0)

    sum_data = groupDataDoanhSo.select_dtypes(include=['number']).sum()
    total_df = pd.DataFrame(sum_data).T
    # Thêm các cột không phải là số vào dòng tổng
    for col in groupDataDoanhSo.columns:
        if col not in total_df.columns:
            total_df[col] = ''  
    # Đặt lại thứ tự các cột để khớp với DataFrame gốc
    total_df = total_df[groupDataDoanhSo.columns]
    total_df["Nhân viên"] = "Tổng"
    # Nối dòng tổng với DataFrame gốc
    groupDataDoanhSo = pd.concat([groupDataDoanhSo, total_df])

    return groupDataDoanhSo

def get_data_report_chi_tieu(location):
    data = get_data_chi_tieu(location,["Ngày chi", "Phân loại", "Lượng chi"])
    data = filter_date(data, "Ngày chi")

    totalChiTieu = data["Lượng chi"].sum()
    data = data[["Phân loại", "Lượng chi"]].groupby("Phân loại").sum().reset_index()
    blank = totalChiTieu - data["Lượng chi"].sum()
    blankRow = pd.DataFrame({'Phân loại': ['Blank'], 'Lượng chi': blank})
    totalRow = pd.DataFrame({'Phân loại': ['Tổng cộng'], 'Lượng chi': totalChiTieu})
    data = pd.concat([data, blankRow], ignore_index=True)
    data = pd.concat([data, totalRow], ignore_index=True)
    data = data.fillna(0)
    return data

def get_data_luong_tong_hop():
    # list all excel file luong nhân viên
    list_of_report_ca_nhan_path = []
    for root, dir, files in os.walk(report_folder):
        for file in files:
            if file.endswith(".xlsx") and ("NV-" in file) and ("Tổng hợp lương nhân viên" not in file):
                list_of_report_ca_nhan_path.append(os.path.join(root, file))
    if list_of_report_ca_nhan_path:
        ret_data = pd.DataFrame()
        for report_ca_nhan_path in list_of_report_ca_nhan_path:
            file_name = os.path.basename(report_ca_nhan_path)
            file_name_part = file_name.split(" ")
            ma_nhan_vien = file_name_part[0]
            ten_nhan_vien = ' '.join(file_name_part[1:-1])
            data_luong = pd.read_excel(report_ca_nhan_path, sheet_name="Lương")

            row_data = {
                "Mã nhân viên" : [ma_nhan_vien], 
                "Tên nhân viên" : [ten_nhan_vien], 
            }
            for location in location_list:
                # Lấy lương thực nhận
                tong_luong_thuc_nhan = data_luong.set_index("Danh mục lương").transpose()[f"Tổng lương tại {location}"]
                if len(tong_luong_thuc_nhan):
                    tong_luong_thuc_nhan = tong_luong_thuc_nhan.values[0]
                else:
                    tong_luong_thuc_nhan = 0
                row_data[f"Tổng lương thực nhận tại {location}"] = [tong_luong_thuc_nhan]
                # Lương ứng
                if location != "HỆ THỐNG":
                    luong_ung = data_luong.set_index("Danh mục lương").transpose()[f"Ứng lương tại {location}"]
                    if len(luong_ung):
                        luong_ung = luong_ung.values[0]
                    else:
                        luong_ung = 0
                    row_data[f"Ứng lương tại {location}"] = [-luong_ung]
                else:
                    luong_ung = 0
                # Tổng lương
                row_data[f"Tổng lương tại {location}"] = [-luong_ung + tong_luong_thuc_nhan]

            tong_ung_luong = 0
            for location in location_list:
                if location != "HỆ THỐNG":
                    tong_ung_luong += float(row_data[f"Ứng lương tại {location}"][0])
            row_data["Ứng lương"] = [tong_ung_luong]

            df_row_data = pd.DataFrame(row_data, columns=list(row_data.keys()))
            ret_data = pd.concat([ret_data, df_row_data])

        # Tạo dòng tính tổng lương
        row_total = {
            "Mã nhân viên" : "Tổng", 
            "Tên nhân viên" : [""], 
        }
        for col in row_data.keys():
            if (col != "Mã nhân viên") and (col != "Tên nhân viên"):
                tong = ret_data[col].sum()
                row_total[col] = [tong]
        df_row_total = pd.DataFrame(row_total, columns=list(row_data.keys()))
        ret_data = pd.concat([ret_data, df_row_total]) 

        # Chuyển cột lương HỆ THỐNG xuống cuối
        ret_data["Tổng lương tại HỆ THỐNG"] = ret_data.pop("Tổng lương tại HỆ THỐNG")
        return ret_data 
    else:
        return None

def get_data_loi_nhuan(location, data_doanh_thu, data_thu_no, data_chi_tieu, data_luong):
    data_thu_no = filter_date_don_no(data_thu_no, "Ngày thực hiện")
    if location == "HỆ THỐNG":
        don_gia = data_doanh_thu["Đơn giá"].sum()/2
        doanh_thu_sale = data_doanh_thu["Đã thanh toán"].sum()/2
        doanh_thu_thu_no = data_thu_no["Lượng thu"].sum()/2
        chi_tieu = data_chi_tieu["Lượng chi"].sum()/2
    else:
        don_gia = data_doanh_thu[data_doanh_thu["Cơ sở"] == location]["Đơn giá"].sum()
        doanh_thu_sale = data_doanh_thu[data_doanh_thu["Cơ sở"] == location]["Đã thanh toán"].sum()
        doanh_thu_thu_no = data_thu_no[data_thu_no["Cơ sở"] == location]["Lượng thu"].sum()
        chi_tieu = data_chi_tieu[data_chi_tieu["Cơ sở"] == location]["Lượng chi"].sum()
    luong = data_luong[f"Tổng lương thực nhận tại {location}"].sum()/2
    loi_nhuan = doanh_thu_sale + doanh_thu_thu_no - (chi_tieu + luong)
    row = {
        "Cơ sở" : [location],
        "Tổng đơn giá" : [don_gia],
        "Đã thanh toán" : [doanh_thu_sale],
        "Tỉ lệ thanh toán" : [doanh_thu_sale/don_gia],
        "Tỉ lệ nợ" : [1-doanh_thu_sale/don_gia],
        "Thu nợ" : [doanh_thu_thu_no],
        "Tổng doanh thu" : [doanh_thu_sale + doanh_thu_thu_no],
        "Chi tiêu" : [chi_tieu],
        "Quỹ lương" : [luong],
        "Tổng chi phí" : [chi_tieu + luong],
        "Lợi nhuận" : [loi_nhuan],
        "Tỉ lệ lợi nhuận" : [loi_nhuan/(doanh_thu_sale + doanh_thu_thu_no)]
    }
    df_row = pd.DataFrame(row, columns=list(row.keys()))
    return df_row

def create_report_co_so(path, location):
    excel_file_path = os.path.join(path, f"{location} {month} - {year}.xlsx")
    # Kiểm tra xem file Excel đã tồn tại hay chưa
    if os.path.exists(excel_file_path):
        # Nếu đã tồn tại, xóa file cũ đi
        try:
            os.remove(excel_file_path)
            print(f"Đã xóa file Excel cũ '{excel_file_path}'")
        except Exception as e:
            print(f"Lỗi khi xóa file Excel cũ: {e}")
    # Tạo workbook mới
    wb = Workbook()
    # Tạo report về Doanh số
    ws1 = wb.active
    ws1.title = 'CHI TIẾT DOANH THU'
    data_doanh_thu = get_data_chi_tiet_doanh_thu(location)
    writeDataframeToSheet(ws1, data_doanh_thu)
    # Tại report chi tiết về thu nợ
    ws2 = wb.create_sheet(title="CHI TIẾT VỀ THU NỢ")
    data_thu_no = get_data_chi_tiet_thu_no(location)
    writeDataframeToSheet(ws2, data_thu_no)
    # Tạo report chi tiết về chi tiêu
    ws3 = wb.create_sheet(title="CHI TIẾT CHI TIÊU")
    data_chi_tieu = get_data_chi_tiet_chi_tieu(location)
    writeDataframeToSheet(ws3, data_chi_tieu)
    # Tạo report về doanh số cá nhân
    ws4 = wb.create_sheet(title='DOANH SỐ CÁ NHÂN')
    writeDataframeToSheet(ws4, get_data_report_doanh_so(location))
    # Tạo report về chi tiêu
    ws5 = wb.create_sheet(title='CHI TIÊU TỔNG HỢP')
    writeDataframeToSheet(ws5, get_data_report_chi_tieu(location))
    # Tạo report về lũy kế ngày
    ws6 = wb.create_sheet(title="LŨY KẾ NGÀY")
    data = get_data_cho_luy_ke(location)
    data = filter_date(data, "Ngày")
    total_row = data.sum()
    total_df = pd.DataFrame(total_row).T
    total_df["Ngày"] = "Tổng"
    data = pd.concat([data, total_df], ignore_index=True)
    data["Lũy kế ngày"] = data["Thanh toán lần đầu"] + data["Thu nợ"] - data["Lượng chi"]
    writeDataframeToSheet(ws6, data)
    # Tạo report về lương tại các cơ sở
    data_luong = get_data_luong_tong_hop()
    ws7 = wb.create_sheet(title="QUỸ LƯƠNG")
    if location != "HỆ THỐNG":
        writeDataframeToSheet(ws7, data_luong[["Mã nhân viên", "Tên nhân viên", f"Tổng lương tại {location}", f"Ứng lương tại {location}" ,f"Tổng lương thực nhận tại {location}"]])
    else:
        col = ["Mã nhân viên", "Tên nhân viên"]
        for location in location_list:
            if location != "HỆ THỐNG":
                col.append(f"Tổng lương tại {location}")
        col.append("Ứng lương")
        col.append("Tổng lương thực nhận tại HỆ THỐNG")
        writeDataframeToSheet(ws7, data_luong[col])
    # Tạo report lợi nhuận
    ws8 = wb.create_sheet(title="LỢI NHUẬN")
    if location != "HỆ THỐNG":
        writeDataframeToSheet(ws8, get_data_loi_nhuan(location, data_doanh_thu, data_thu_no, data_chi_tieu, data_luong))
    else:
        loi_nhuan_he_thong = get_data_loi_nhuan(location_list[0], data_doanh_thu, data_thu_no, data_chi_tieu, data_luong)
        for i in range(1, len(location_list)):
            loi_nhuan_he_thong = pd.concat([loi_nhuan_he_thong,get_data_loi_nhuan(location_list[i], data_doanh_thu, data_thu_no, data_chi_tieu, data_luong)])
        
        writeDataframeToSheet(ws8, loi_nhuan_he_thong)


    # Lưu workbook vào file Excel
    try:
        wb.save(excel_file_path)
        print(f"Đã tạo file Excel mới '{excel_file_path}' thành công")
    except Exception as e:
        print(f"Lỗi khi tạo file Excel mới: {e}")


            