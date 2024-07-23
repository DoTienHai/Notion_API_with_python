import os
from openpyxl import Workbook
import pandas as pd
from Config import *
from UpdateLuyKe import get_data_cho_luy_ke
from CreateReportCaNhan import filter_date, add_total_row

start_date = '2024-07-01'
end_date = '2024-07-30'

def get_data_chi_tiet_doanh_thu(location = ""):
    data = get_data_doanh_thu(location, ["ALL"])
    data = data[["Tiền tố", "Mã dịch vụ", "Ngày thực hiện", "Cơ sở","Tên dịch vụ", "Khách hàng",
                 "Nguồn khách", "Sale chính", "Đơn giá gốc", "Sale phụ", "Upsale", "Đơn giá", 
                 "Bác sĩ 1", "Bác sĩ 2", "Thanh toán lần đầu", "Trả sau", "Đã thanh toán", 
                 "Dư nợ", "Phụ phẫu 1", "Phụ phẫu 2", "Công phụ phẫu 1", "Công phụ phẫu 2"]]
    data = filter_date(data, "Ngày thực hiện")
    data = add_total_row(data)
    return data

def get_data_chi_tiet_chi_tieu(location=""):
    data = get_data_chi_tieu(location, ["ALL"])
    data = filter_date(data, "Ngày chi")
    data = data[["Tiền tố", "Mã chi tiêu", "Ngày chi", "Cơ sở", "Phân loại", "Lượng chi"]]
    data = add_total_row(data)
    return data

def get_data_report_doanh_so(location = ""):
    data = get_data_doanh_thu(location,["ALL"])
    if (location):
        data = data[data["Cơ sở"] == location]
    query_string = f"'{start_date}' <= `Ngày thực hiện` <= '{end_date}'"
    data = data.query(query_string)

    groupDataDoanhSo = pd.DataFrame(columns=["Mã nhân viên"])
    # Group data Sale chính
    groupDataSaleChinh = data.groupby("Sale chính")[["Đơn giá gốc"]].sum().reset_index()
    groupDataSaleChinh = groupDataSaleChinh.rename(columns={"Sale chính":"Mã nhân viên", "Đơn giá gốc":"Tổng đơn giá sale vòng 1"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataSaleChinh, on='Mã nhân viên', how='outer')
    # Group data Sale phụ
    groupDataSalePhu = data.groupby("Sale phụ")[["Upsale"]].sum().reset_index()
    groupDataSalePhu = groupDataSalePhu.rename(columns={"Sale phụ":"Mã nhân viên", "Upsale":"Tổng đơn giá vòng upsale"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataSalePhu, on='Mã nhân viên', how='outer')
    # Group data 1 bác sĩ
    groupData1BacSi = data[data["id bác sĩ 2"].isnull()]
    groupData1BacSi = groupData1BacSi.groupby("Bác sĩ 1")[["Đã thanh toán"]].sum().reset_index()
    groupData1BacSi = groupData1BacSi.rename(columns={"Bác sĩ 1":"Mã nhân viên", "Đã thanh toán":"Doanh số đơn 1 bác sĩ"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupData1BacSi, on='Mã nhân viên', how='outer')
    # Group data 2 bác sĩ
    temp = data[~data["id bác sĩ 2"].isnull()]
    groupData2BacSi = temp.groupby("Bác sĩ 1")[["Đã thanh toán"]].sum().reset_index()
    groupData2BacSi = groupData2BacSi.rename(columns={"Bác sĩ 1":"Mã nhân viên", "Đã thanh toán":"A"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupData2BacSi, on='Mã nhân viên', how='outer')
    groupData2BacSi = temp.groupby("Bác sĩ 2")[["Đã thanh toán"]].sum().reset_index()
    groupData2BacSi = groupData2BacSi.rename(columns={"Bác sĩ 2":"Mã nhân viên", "Đã thanh toán":"B"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupData2BacSi, on='Mã nhân viên', how='outer')
    groupDataDoanhSo["Doanh số đơn 2 bác sĩ"] = groupDataDoanhSo["A"] + groupDataDoanhSo["B"]
    # group data phụ phẫu 1
    groupDataCountPhuPhau1 = data["Phụ phẫu 1"].value_counts().reset_index()
    groupDataCountPhuPhau1 = groupDataCountPhuPhau1.rename(columns={"Phụ phẫu 1":"Mã nhân viên", "count":"Số lần phụ phẫu 1"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCountPhuPhau1, on='Mã nhân viên', how='outer')
    # group data công phụ phẫu 1
    groupDataCongPhuPhau1 = data.groupby("Phụ phẫu 1")[["Công phụ phẫu 1"]].sum().reset_index()
    groupDataCongPhuPhau1 = groupDataCongPhuPhau1.rename(columns={"Phụ phẫu 1":"Mã nhân viên"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCongPhuPhau1, on='Mã nhân viên', how='outer')
    # group data phụ phẫu 2
    groupDataCountPhuPhau2 = data["Phụ phẫu 2"].value_counts().reset_index()
    groupDataCountPhuPhau2 = groupDataCountPhuPhau2.rename(columns={"Phụ phẫu 2":"Mã nhân viên", "count":"Số lần phụ phẫu 2"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCountPhuPhau2, on='Mã nhân viên', how='outer')
    # group data công phụ phẫu 2
    groupDataCongPhuPhau2 = data.groupby("Phụ phẫu 2")[["Công phụ phẫu 2"]].sum().reset_index()
    groupDataCongPhuPhau2 = groupDataCongPhuPhau2.rename(columns={"Phụ phẫu 2":"Mã nhân viên"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataCongPhuPhau2, on='Mã nhân viên', how='outer')
    #group data thu nơ
    groupDataThuNo = get_data_thu_no(location, ["ALL"])
    if (location):
        groupDataThuNo = groupDataThuNo[groupDataThuNo["Cơ sở"] == location]
    query_string = f"'{start_date}' <= `Ngày thu` <= '{end_date}'"
    groupDataThuNo = groupDataThuNo.query(query_string)
    groupDataThuNo = groupDataThuNo.groupby("Sale chính")[["Lượng thu"]].sum().reset_index()
    groupDataThuNo = groupDataThuNo.rename(columns={"Sale chính":"Mã nhân viên", "Lượng thu":"Doanh số thu nợ"})
    groupDataDoanhSo = pd.merge(groupDataDoanhSo, groupDataThuNo, on='Mã nhân viên', how='outer')
    groupDataDoanhSo = groupDataDoanhSo.drop(columns=["A","B"])
    groupDataDoanhSo = groupDataDoanhSo.fillna(0)

    sum_data = groupDataDoanhSo.select_dtypes(include=['number']).sum()
    total_df = pd.DataFrame(sum_data).T
    # Thêm các cột không phải là số vào dòng tổng
    for col in groupDataDoanhSo.columns:
        if col not in total_df.columns:
            total_df[col] = ''  
    # Đặt lại thứ tự các cột để khớp với DataFrame gốc
    total_df = total_df[groupDataDoanhSo.columns]
    total_df["Mã nhân viên"] = "Tổng"
    # Nối dòng tổng với DataFrame gốc
    groupDataDoanhSo = pd.concat([groupDataDoanhSo, total_df])
    return groupDataDoanhSo

def get_data_report_chi_tieu(location = ""):
    data = get_data_chi_tieu(location,["Ngày chi", "Phân loại", "Lượng chi"])
    query_string = f"'{start_date}' <= `Ngày chi` <= '{end_date}'"
    data = data.query(query_string)
    totalChiTieu = data["Lượng chi"].sum()
    data = data[["Phân loại", "Lượng chi"]].groupby("Phân loại").sum().reset_index()
    blank = totalChiTieu - data["Lượng chi"].sum()
    blankRow = pd.DataFrame({'Phân loại': ['Blank'], 'Lượng chi': blank})
    totalRow = pd.DataFrame({'Phân loại': ['Tổng cộng'], 'Lượng chi': totalChiTieu})
    data = pd.concat([data, blankRow], ignore_index=True)
    data = pd.concat([data, totalRow], ignore_index=True)
    data = data.fillna(0)
    return data

def createReportLocation(location = ""):
    if(location != ""):
        excel_file_path = os.path.join(report_co_so_folder, f"{location} {start_date.replace('/', '_')} - {end_date.replace('/', '_')}.xlsx")
        # Kiểm tra xem file Excel đã tồn tại hay chưa
        if os.path.exists(excel_file_path):
            # Nếu đã tồn tại, xóa file cũ đi
            try:
                os.remove(excel_file_path)
                print(f"Đã xóa file Excel cũ '{excel_file_path}'")
            except Exception as e:
                print(f"Lỗi khi xóa file Excel cũ: {e}")
        # Tạo workbook mới
        wb = Workbook()
        # Tạo report về Doanh số
        ws1 = wb.active
        ws1.title = 'CHI TIẾT DOANH THU'
        writeDataframeToSheet(ws1, get_data_chi_tiet_doanh_thu(location))
        # Tạo report chi tiết về chi tiêu
        ws2 = wb.create_sheet(title="CHI TIẾT CHI TIÊU")
        writeDataframeToSheet(ws2, get_data_chi_tiet_chi_tieu(location))
        # Tạo report về doanh số cá nhân
        ws3 = wb.create_sheet(title='DOANH SỐ CÁ NHÂN')
        writeDataframeToSheet(ws3, get_data_report_doanh_so(location))
        # Tạo report về chi tiêu
        ws4 = wb.create_sheet(title='CHI TIÊU TỔNG HỢP')
        writeDataframeToSheet(ws4, get_data_report_chi_tieu(location))
        # Tạo report về lũy kế ngày
        ws5 = wb.create_sheet(title="LŨY KẾ NGÀY")
        query_string = f"'{start_date}' <= `Ngày` <= '{end_date}'"
        data = get_data_cho_luy_ke(location).query(query_string)
        data = filter_date(data, "Ngày")
        total_row = data.sum()
        total_df = pd.DataFrame(total_row).T
        total_df["Ngày"] = "Tổng"
        data = pd.concat([data, total_df], ignore_index=True)
        data["Lũy kế ngày"] = data["Thanh toán lần đầu"] + data["Thu nợ"] - data["Lượng chi"]
        writeDataframeToSheet(ws5, data)

        # Lưu workbook vào file Excel
        try:
            wb.save(excel_file_path)
            print(f"Đã tạo file Excel mới '{excel_file_path}' thành công")
        except Exception as e:
            print(f"Lỗi khi tạo file Excel mới: {e}")
    else:
        print("Sai tên cơ sở! Không thể tạo report cho cở sở!") 

def createReportSystem():
        excel_file_path = os.path.join(report_co_so_folder, f"HỆ THỐNG {start_date.replace('/', '_')} - {end_date.replace('/', '_')}.xlsx")
        # Kiểm tra xem file Excel đã tồn tại hay chưa
        if os.path.exists(excel_file_path):
            # Nếu đã tồn tại, xóa file cũ đi
            try:
                os.remove(excel_file_path)
                print(f"Đã xóa file Excel cũ '{excel_file_path}'")
            except Exception as e:
                print(f"Lỗi khi xóa file Excel cũ: {e}")
        # Tạo workbook mới
        wb = Workbook()
        # Tạo report về Doanh số
        ws1 = wb.active
        ws1.title = 'CHI TIẾT DOANH THU'
        writeDataframeToSheet(ws1, get_data_chi_tiet_doanh_thu())
        # Tạo report chi tiết về chi tiêu
        ws2 = wb.create_sheet(title="CHI TIẾT CHI TIÊU")
        writeDataframeToSheet(ws2, get_data_chi_tiet_chi_tieu())
        # Tạo report về doanh số cá nhân
        ws3 = wb.create_sheet(title='DOANH SỐ CÁ NHÂN')
        writeDataframeToSheet(ws3, get_data_report_doanh_so())
        # Tạo report về chi tiêu
        ws4 = wb.create_sheet(title='CHI TIÊU TỔNG HỢP')
        writeDataframeToSheet(ws4, get_data_report_chi_tieu())
        # Tạo report về lũy kế ngày
        ws5 = wb.create_sheet(title="LŨY KẾ NGÀY")
        query_string = f"'{start_date}' <= `Ngày` <= '{end_date}'"
        data = get_data_cho_luy_ke().query(query_string)
        data = filter_date(data, "Ngày")
        total_row = data.sum()
        total_df = pd.DataFrame(total_row).T
        total_df["Ngày"] = "Tổng"
        data = pd.concat([data, total_df], ignore_index=True)
        writeDataframeToSheet(ws5, data)
   
        # Lưu workbook vào file Excel
        try:
            wb.save(excel_file_path)
            print(f"Đã tạo file Excel mới '{excel_file_path}' thành công")
        except Exception as e:
            print(f"Lỗi khi tạo file Excel mới: {e}")


def create_report_co_so():
    createReportSystem()
    for location in vn_locations:
        createReportLocation(location)

# create_report_co_so()

            